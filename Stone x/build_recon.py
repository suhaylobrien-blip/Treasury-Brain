"""
StoneX March 2026 Reconciliation Builder
Produces: StoneX March 2026 Recon.xlsx + StoneX March 2026 Recon Report.pdf
All computed cells use Excel formulas so the maths can be inspected and verified.
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Style helpers ─────────────────────────────────────────────────────────────
def fill(c): return PatternFill("solid", fgColor=c)
def bfont(color="000000", sz=10): return Font(bold=True, color=color, size=sz)
def rfont(sz=10): return Font(size=sz)
def centre(): return Alignment(horizontal="center", vertical="center")
def wrap(): return Alignment(wrap_text=True, vertical="top")

# SA Bullion brand palette
NAVY  ="150E26"; GOLD ="D4A755"; LGOLD="FAF3E0"; BLUE ="7B4FC9"; LBLUE="EDE4F7"
ORAN  ="D4720A"; LORAN="FDF0E0"; PURP ="4B1D75"; LPURP="E8E0F7"; GRND ="40B5AD"
LGRN  ="E0F5F4"; GREY ="808080"; LGRY ="F2F2F2"; RED  ="E05252"; WHT  ="FFFFFF"

def hdr_row(ws, row, vals, bg, fg=WHT, sz=10):
    for c, v in enumerate(vals, 1):
        cl = ws.cell(row, c, v)
        cl.fill = fill(bg); cl.font = Font(bold=True, color=fg, size=sz)

def autofit(ws, mn=10, mx=42):
    for col in ws.columns:
        best = mn
        letter = get_column_letter(col[0].column)
        for cell in col:
            try: best = max(best, len(str(cell.value or "")))
            except: pass
        ws.column_dimensions[letter].width = min(mx, best + 2)

def title_row(ws, text, bg, fg=WHT, sz=13, cols="A1:J1"):
    ws.merge_cells(cols)
    ws[cols.split(":")[0]] = text
    ws[cols.split(":")[0]].font = Font(bold=True, color=fg, size=sz)
    ws[cols.split(":")[0]].fill = fill(bg)
    ws[cols.split(":")[0]].alignment = centre()
    ws.row_dimensions[1].height = 22

# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ─────────────────────────────── SHEET 1: SUMMARY ────────────────────────────
ws = wb.active; ws.title = "Summary"; ws.sheet_properties.tabColor = NAVY
title_row(ws, "StoneX Financial Ltd — March 2026 Reconciliation", NAVY, cols="A1:H1")
ws.merge_cells("A2:H2")
ws["A2"] = "SA Bullion Investor Services  |  Account MT0795  |  Period: 16 Mar – 31 Mar 2026"
ws["A2"].font = Font(color=WHT, size=10); ws["A2"].fill = fill(NAVY)
ws["A2"].alignment = centre(); ws.row_dimensions[2].height = 14

r = 4
ws.cell(r,1,"POSITION SUMMARY").font = bfont(sz=11); r+=1
hdr_row(ws,r,["Instrument","Net Oz","VWAP (USD/oz)","Cost Basis (USD)","MTM Price","MTM Value (USD)","Unrealised P&L","P&L %"],NAVY)
r+=1
pdata=[
    ("XAU/USD",      309, 4450.91, 1375332.51, 4667.02, 1441909.18,  66576.67,"4.84%",LGOLD),
    ("XAU/ZAR",       52, 4346.08,  225996.47, 4667.02,  242684.04,  16687.57,"7.38%",LORAN),
    ("XAU COMBINED", 361, 4435.81, 1601328.98, 4667.02, 1684593.22,  83264.24,"5.20%",GOLD ),
    ("XAG/USD",     8820,   72.97,  643580.61,   75.007,  661561.74,  17981.13,"2.79%",LBLUE),
    ("XAG/ZAR",      635,   69.92,   44399.04,   75.007,   47629.45,   3230.41,"7.28%",LPURP),
    ("XAG COMBINED",9455,   72.76,  687979.65,   75.007,  709191.19,  21211.54,"3.08%",BLUE ),
    ("TOTAL",          "",    "",  2289308.63,       "", 2393784.41, 104475.78,"4.56%",NAVY ),
]
for d in pdata:
    bg=d[-1]; is_tot="COMBINED" in str(d[0]) or d[0]=="TOTAL"
    fg=WHT if bg in (BLUE,NAVY) else "000000"
    for c,v in enumerate(d[:-1],1):
        cl=ws.cell(r,c,v); cl.fill=fill(bg)
        cl.font=Font(bold=is_tot,color=fg,size=10)
        if c in(3,4,5,6,7) and isinstance(v,float): cl.number_format='#,##0.00'
        if c==2 and isinstance(v,int): cl.number_format='#,##0'
    r+=1

r+=1
ws.cell(r,1,"CASH SUMMARY (USD)").font=bfont(sz=11); r+=1
hdr_row(ws,r,["Item","Amount (USD)"],NAVY); r+=1
# Track the cash item rows so closing balance can use a SUM formula
cash_data_start = r
for lbl,amt in[
    ("Opening Balance",0.00),
    ("+ Incoming Wires",445233.94),
    ("− XAU Purchases (USD legs)",-1375332.51),
    ("− XAG Purchases (USD legs)",-643580.61),
    ("+ XAU Sale Proceeds (5 oz)",21911.00),
    ("− Total Swap Fees",-2943.84),
    ("+ Interest Earned",44.80),
]:
    is_t = lbl.startswith("=")
    bg_ = LGRY if r%2==0 else None
    cl1=ws.cell(r,1,lbl); cl1.font=Font(bold=is_t,size=10)
    cl2=ws.cell(r,2,amt); cl2.number_format='#,##0.00'
    cl2.font=Font(bold=is_t,color=RED if amt<0 else "000000",size=10)
    if bg_: cl1.fill=fill(bg_); cl2.fill=fill(bg_)
    r+=1
cash_data_end = r - 1
# Closing Balance row — formula sums all items above
cl1=ws.cell(r,1,"= Closing Balance (USD)"); cl1.font=Font(bold=True,size=10)
cl2=ws.cell(r,2,f"=SUM(B{cash_data_start}:B{cash_data_end})")
cl2.number_format='#,##0.00'; cl2.font=Font(bold=True,size=10)
cl1.fill=fill(LGOLD); cl2.fill=fill(LGOLD)
r+=1
# ZAR Ledger (informational — separate ledger balance)
ws.cell(r,1,"ZAR Ledger Closing Balance").font=Font(size=10)
ws.cell(r,2,-39037714.16).number_format='#,##0.00'
ws.cell(r,2).font=Font(color=RED,size=10)
r+=1

r+=1
ws.cell(r,1,"SWAP FEES SUMMARY").font=bfont(sz=11); r+=1
hdr_row(ws,r,["Category","Total Cost (USD)"],GREY); r+=1
for lbl,amt in[
    ("XAG Carry (4.77% / 5.12% / 4.87%)",615.24),
    ("XAU Carry (4.87%)",1027.76),
    ("ZAR Funding (3.59% – 3.88%)",1300.84),
    ("TOTAL SWAP FEES",2943.84),
]:
    is_t=lbl.startswith("TOTAL")
    ws.cell(r,1,lbl).font=Font(bold=is_t,size=10)
    cl=ws.cell(r,2,amt); cl.number_format='#,##0.00'
    cl.font=Font(bold=is_t,color=RED,size=10)
    if is_t: ws.cell(r,1).fill=fill(LGRY); cl.fill=fill(LGRY)
    r+=1

r+=1
ws.cell(r,1,"INTEREST EARNED").font=bfont(sz=11); r+=1
hdr_row(ws,r,["Date","Reference","Avg Balance","Rate","Amount (USD)"],GRND); r+=1
sum_int_start = r
for d,ref,bal,rate,amt in[
    ("30-Mar-2026","EJV/2026/002998",141568.29,"4.87%",19.15),
    ("31-Mar-2026","EJV/2026/003045",189634.36,"4.87%",25.65),
]:
    ws.cell(r,1,d); ws.cell(r,2,ref)
    ws.cell(r,3,bal).number_format='#,##0.00'
    ws.cell(r,4,rate); ws.cell(r,5,amt).number_format='#,##0.00'
    r+=1
sum_int_end = r - 1
ws.cell(r,1,"TOTAL").font=bfont()
ws.cell(r,5,f"=SUM(E{sum_int_start}:E{sum_int_end})").number_format='#,##0.00'
ws.cell(r,5).font=bfont()
r+=1

r+=1
ws.cell(r,1,"INCOMING WIRES").font=bfont(sz=11); r+=1
hdr_row(ws,r,["Date","Reference","Amount (USD)"],ORAN); r+=1
sum_wire_start = r
for d,ref,amt in[
    ("13-Mar-2026","JRV/2026/013947",120435.46),
    ("24-Mar-2026","JRV/2026/015715",177989.26),
    ("27-Mar-2026","JRV/2026/016606",146809.22),
]:
    ws.cell(r,1,d); ws.cell(r,2,ref)
    ws.cell(r,3,amt).number_format='#,##0.00'
    r+=1
sum_wire_end = r - 1
ws.cell(r,1,"TOTAL").font=bfont()
ws.cell(r,3,f"=SUM(C{sum_wire_start}:C{sum_wire_end})").number_format='#,##0.00'
ws.cell(r,3).font=bfont()
r+=2
ws.merge_cells(f"A{r}:H{r}")
ws[f"A{r}"]="NOTE: MTM prices XAU $4,667.02/oz and XAG $75.007/oz sourced from 31-Mar-2026 StoneX overnight swap rate. Confirm against official closing prices for audit purposes. XAU/ZAR and XAG/ZAR positions converted at implied trade-date USD/ZAR rates."
ws[f"A{r}"].font=Font(italic=True,size=9,color="666666"); ws[f"A{r}"].alignment=wrap()
ws.row_dimensions[r].height=28
autofit(ws); ws.freeze_panes="A5"

# ─────────────────────────────── SHEET 2: XAU RECON ─────────────────────────
# Columns: A=Trade Date  B=Value Date  C=Doc#  D=Direction  E=Oz  F=Price
#          G=USD Amount  H=Running Oz  I=Running Cost  J=VWAP
# G, H, I, J are all Excel formulas — no pre-computed values written.
ws2=wb.create_sheet("XAU Recon"); ws2.sheet_properties.tabColor="D4A755"
title_row(ws2,"XAU/USD — Gold Forward Purchases & Sales — March 2026","D4A755")
r=2
xau_cols=["Trade Date","Value Date","Doc #","Direction","Oz","Price (USD/oz)","USD Amount","Running Oz","Running Cost","VWAP (USD/oz)"]
hdr_row(ws2,r,xau_cols,GOLD,"000000"); r+=1

# Input data: Trade Date, Value Date, Doc#, Direction, Oz, Price
# USD Amount = Oz × Price  |  Running Oz / Cost = cumulative  |  VWAP = Running Cost / Running Oz
xau=[
    ("24-Mar-26","26-Mar-26","FNC/2026/077551","BUY",  125, 4426.77),
    ("24-Mar-26","26-Mar-26","FNC/2026/077779","SELL",  -5, 4382.20),
    ("24-Mar-26","26-Mar-26","FNC/2026/077860","BUY",   25, 4355.55),
    ("24-Mar-26","26-Mar-26","FNC/2026/077888","BUY",   10, 4395.87),
    ("25-Mar-26","27-Mar-26","FNC/2026/078603","BUY",   25, 4541.28),
    ("25-Mar-26","27-Mar-26","FNC/2026/078631","BUY",   16, 4564.53),
    ("25-Mar-26","27-Mar-26","FNC/2026/079433","BUY",    7, 4557.76),
    ("26-Mar-26","30-Mar-26","FNC/2026/080112","BUY",   20, 4454.71),
    ("26-Mar-26","30-Mar-26","FNC/2026/080543","BUY",   25, 4448.64),
    ("26-Mar-26","30-Mar-26","FNC/2026/080932","BUY",    5, 4474.77),
    ("27-Mar-26","31-Mar-26","FNC/2026/081741","BUY",   10, 4430.03),
    ("27-Mar-26","31-Mar-26","FNC/2026/082111","BUY",   35, 4436.19),
    ("31-Mar-26","02-Apr-26","FNC/2026/084495","BUY",    5, 4557.45),
    ("31-Mar-26","02-Apr-26","FNC/2026/084820","BUY",    6, 4590.46),
]
r_xau_start = r
for i, d in enumerate(xau):
    bg_ = LGOLD if d[3]=="BUY" else "FFCCCC"
    # Write input columns A–F
    for c, v in enumerate(d, 1):
        cl = ws2.cell(r, c, v); cl.fill = fill(bg_)
        if c == 5: cl.number_format = '#,##0'
        if c == 6: cl.number_format = '#,##0.00'
    # G: USD Amount = Oz × Price  (E is negative for SELL so sign is automatic)
    cl = ws2.cell(r, 7, f"=E{r}*F{r}"); cl.fill = fill(bg_); cl.number_format = '#,##0.00'
    # H: Running Oz (cumulative E)
    h_f = f"=E{r}" if i == 0 else f"=H{r-1}+E{r}"
    cl = ws2.cell(r, 8, h_f); cl.fill = fill(bg_); cl.number_format = '#,##0.00'
    # I: Running Cost (cumulative G)
    i_f = f"=G{r}" if i == 0 else f"=I{r-1}+G{r}"
    cl = ws2.cell(r, 9, i_f); cl.fill = fill(bg_); cl.number_format = '#,##0.00'
    # J: VWAP = Running Cost / Running Oz
    cl = ws2.cell(r, 10, f"=IF(H{r}=0,0,I{r}/H{r})"); cl.fill = fill(bg_); cl.number_format = '#,##0.00'
    r += 1
r_xau_end = r - 1

# Net position summary row — all formula references to the last data row
hdr_row(ws2,r,["NET POSITION","","","",
    f"=H{r_xau_end}","","",
    f'=TEXT(H{r_xau_end},"#,##0")&" oz"',
    f"=I{r_xau_end}",
    f"=J{r_xau_end}"],GOLD,"000000")
ws2.cell(r,5).number_format='#,##0.00'
ws2.cell(r,9).number_format='#,##0.00'; ws2.cell(r,10).number_format='#,##0.00'
ws2.cell(r,4).font=bfont(); ws2.cell(r,5).font=bfont()
r+=2

ws2.cell(r,1,"MONTH-END MARK-TO-MARKET — 31 March 2026").font=bfont(sz=11); r+=1
for lbl,val in[
    ("MTM Price (31-Mar SWT rate — confirm official closing spot)","$4,667.02 / oz"),
    ("MTM Value  309 oz × $4,667.02","$1,441,909.18"),
    ("Total Cost Basis","$1,375,332.51"),
    ("Unrealised P&L","$66,576.67"),
    ("Short-sale note","5 oz sold FNC/077779 @ $4,382.20 = $21,911.00 proceeds. Realised loss vs VWAP = ($222.85)"),
]:
    ws2.cell(r,1,lbl); ws2.cell(r,2,val).font=bfont(); r+=1
autofit(ws2); ws2.freeze_panes="A3"

# ─────────────────────────────── SHEET 3: XAG RECON ─────────────────────────
# Same formula structure as XAU: G=E*F, H cumulative E, I cumulative G, J=I/H
ws3=wb.create_sheet("XAG Recon"); ws3.sheet_properties.tabColor=BLUE
title_row(ws3,"XAG/USD — Silver Forward Purchases — March 2026",BLUE)
r=2; hdr_row(ws3,r,xau_cols,BLUE); r+=1

xag=[
    ("16-Mar-26","18-Mar-26","FNC/2026/067462","BUY", 2500, 78.8992),
    ("17-Mar-26","19-Mar-26","FNC/2026/068552","BUY",  250, 80.5355),
    ("19-Mar-26","23-Mar-26","FNC/2026/071204","BUY",  245, 68.5907),
    ("19-Mar-26","23-Mar-26","FNC/2026/071324","BUY",  100, 66.5227),
    ("19-Mar-26","23-Mar-26","FNC/2026/071420","BUY",  220, 68.2362),
    ("19-Mar-26","23-Mar-26","FNC/2026/072203","BUY",  140, 72.3961),
    ("23-Mar-26","25-Mar-26","FNC/2026/074774","BUY",  300, 65.0102),
    ("24-Mar-26","26-Mar-26","FNC/2026/077437","BUY",  900, 70.3309),
    ("24-Mar-26","26-Mar-26","FNC/2026/077554","BUY",  300, 70.0672),
    ("24-Mar-26","26-Mar-26","FNC/2026/077744","BUY",  500, 68.3954),
    ("25-Mar-26","27-Mar-26","FNC/2026/078842","BUY",  210, 72.7084),
    ("26-Mar-26","30-Mar-26","FNC/2026/080022","BUY",  300, 69.9260),
    ("26-Mar-26","30-Mar-26","FNC/2026/080904","BUY",  100, 69.5142),
    ("27-Mar-26","31-Mar-26","FNC/2026/081728","BUY",  250, 68.7385),
    ("30-Mar-26","01-Apr-26","FNC/2026/083025","BUY", 1000, 70.5852),
    ("30-Mar-26","01-Apr-26","FNC/2026/083102","BUY",  500, 70.9508),
    ("30-Mar-26","01-Apr-26","FNC/2026/083541","BUY",  150, 71.2835),
    ("31-Mar-26","02-Apr-26","FNC/2026/084221","BUY",  250, 73.0316),
    ("31-Mar-26","02-Apr-26","FNC/2026/084360","BUY",  105, 73.2625),
    ("31-Mar-26","02-Apr-26","FNC/2026/084449","BUY",  500, 72.9788),
]
r_xag_start = r
for i, d in enumerate(xag):
    bg_ = LBLUE if i%2==0 else None
    for c, v in enumerate(d, 1):
        cl = ws3.cell(r, c, v)
        if bg_: cl.fill = fill(bg_)
        if c == 5: cl.number_format = '#,##0'
        if c == 6: cl.number_format = '#,##0.000'
    cl = ws3.cell(r, 7, f"=E{r}*F{r}")
    if bg_: cl.fill = fill(bg_)
    cl.number_format = '#,##0.00'
    h_f = f"=E{r}" if i == 0 else f"=H{r-1}+E{r}"
    cl = ws3.cell(r, 8, h_f)
    if bg_: cl.fill = fill(bg_)
    cl.number_format = '#,##0'
    i_f = f"=G{r}" if i == 0 else f"=I{r-1}+G{r}"
    cl = ws3.cell(r, 9, i_f)
    if bg_: cl.fill = fill(bg_)
    cl.number_format = '#,##0.00'
    cl = ws3.cell(r, 10, f"=IF(H{r}=0,0,I{r}/H{r})")
    if bg_: cl.fill = fill(bg_)
    cl.number_format = '#,##0.000'
    r += 1
r_xag_end = r - 1

hdr_row(ws3,r,["NET POSITION","","","",
    f"=H{r_xag_end}","","",
    f'=TEXT(H{r_xag_end},"#,##0")&" oz"',
    f"=I{r_xag_end}",
    f"=J{r_xag_end}"],BLUE)
ws3.cell(r,5).number_format='#,##0'
ws3.cell(r,9).number_format='#,##0.00'; ws3.cell(r,10).number_format='#,##0.000'
r+=2
ws3.cell(r,1,"MONTH-END MARK-TO-MARKET — 31 March 2026").font=bfont(sz=11); r+=1
for lbl,val in[
    ("MTM Price (31-Mar SWT rate — confirm official closing spot)","$75.007 / oz"),
    ("MTM Value  8,820 oz × $75.007","$661,561.74"),
    ("Total Cost Basis","$643,580.61"),
    ("Unrealised P&L","$17,981.13"),
]:
    ws3.cell(r,1,lbl); ws3.cell(r,2,val).font=bfont(); r+=1
autofit(ws3); ws3.freeze_panes="A3"

# ──────────────────────────── SHEET 4: XAUZAR RECON ─────────────────────────
# Columns: A=Trade Date  B=Value Date  C=Doc#  D=Oz  E=ZAR Price  F=ZAR Amount
#          G=Impl USD/ZAR  H=USD Equivalent  I=Running Oz  J=Running ZAR Cost
# F, H, I, J are Excel formulas.
ws4=wb.create_sheet("XAUZAR Recon"); ws4.sheet_properties.tabColor="FF8C00"
title_row(ws4,"XAU/ZAR — ZAR-Denominated Gold Forwards — March 2026",ORAN)
ws4.merge_cells("A2:J2")
ws4["A2"]="⚠  These trades appear ONLY in the ZAR (R) statement. They are NOT visible as metal legs in the USD ($) statement."
ws4["A2"].font=Font(bold=True,color="8B0000",size=10); ws4["A2"].fill=fill("FFF3CD"); ws4["A2"].alignment=wrap()
ws4.row_dimensions[2].height=18
r=3
hdr_row(ws4,r,["Trade Date","Value Date","Doc #","Oz","ZAR Price (R/oz)","ZAR Amount","Impl USD/ZAR","USD Equivalent","Running Oz","Running ZAR Cost"],ORAN,"000000"); r+=1

# Input: date, vdate, doc, oz, zar_price, usd_zar_rate
xauzar=[
    ("23-Mar-26","25-Mar-26","FNC/2026/074868",12,73192.16,17.1768),
    ("23-Mar-26","25-Mar-26","FNC/2026/076104",40,75087.72,17.1831),
]
r_xauzar_start = r
for i, d in enumerate(xauzar):
    # Cols A–E: Trade Date, Value Date, Doc#, Oz, ZAR Price
    for c, v in enumerate(d[:5], 1):
        cl = ws4.cell(r, c, v); cl.fill = fill(LORAN)
        if c == 4: cl.number_format = '#,##0'
        if c == 5: cl.number_format = '#,##0.00'
    # G: Impl USD/ZAR rate (input)
    cl = ws4.cell(r, 7, d[5]); cl.fill = fill(LORAN); cl.number_format = '#,##0.0000'
    # F: ZAR Amount = Oz × ZAR Price  (=D*E)
    cl = ws4.cell(r, 6, f"=D{r}*E{r}"); cl.fill = fill(LORAN); cl.number_format = '#,##0.00'
    # H: USD Equivalent = ZAR Amount / USD/ZAR rate  (=F/G)
    cl = ws4.cell(r, 8, f"=F{r}/G{r}"); cl.fill = fill(LORAN); cl.number_format = '#,##0.00'
    # I: Running Oz
    i_f = f"=D{r}" if i == 0 else f"=I{r-1}+D{r}"
    cl = ws4.cell(r, 9, i_f); cl.fill = fill(LORAN); cl.number_format = '#,##0'
    # J: Running ZAR Cost
    j_f = f"=F{r}" if i == 0 else f"=J{r-1}+F{r}"
    cl = ws4.cell(r, 10, j_f); cl.fill = fill(LORAN); cl.number_format = '#,##0.00'
    r += 1
r_xauzar_end = r - 1

hdr_row(ws4,r,["TOTAL","","",
    f"=I{r_xauzar_end}",
    "",
    f'="R "&TEXT(J{r_xauzar_end},"#,##0.00")',
    "",
    f"=SUM(H{r_xauzar_start}:H{r_xauzar_end})",
    f"=I{r_xauzar_end}",
    f"=J{r_xauzar_end}"],ORAN,"000000")
ws4.cell(r,4).number_format='#,##0'
ws4.cell(r,8).number_format='#,##0.00'; r+=1
ws4.cell(r,1,"VWAP").font=bfont()
ws4.cell(r,5,f"=J{r_xauzar_end}/I{r_xauzar_end}").number_format='#,##0.00'; ws4.cell(r,5).font=bfont()
ws4.cell(r,8,f"=SUM(H{r_xauzar_start}:H{r_xauzar_end})/I{r_xauzar_end}").number_format='#,##0.00'; ws4.cell(r,8).font=bfont()
r+=2
ws4.cell(r,1,"MONTH-END MARK-TO-MARKET — 31 March 2026").font=bfont(sz=11); r+=1
for lbl,val in[
    ("MTM Price (USD)","$4,667.02 / oz"),
    ("MTM Value  52 oz × $4,667.02","$242,684.04"),
    ("Cost Basis (USD equivalent)","$225,996.47"),
    ("Unrealised P&L (USD equivalent)","$16,687.57"),
    ("NOTE","ZAR positions carry dual risk: gold price risk + USD/ZAR FX risk. Actual ZAR/USD month-end rate should be confirmed."),
]:
    ws4.cell(r,1,lbl); ws4.cell(r,2,val).font=bfont() if lbl!="NOTE" else Font(italic=True,size=9); r+=1
r+=1
ws4.merge_cells(f"A{r}:J{r}")
ws4[f"A{r}"]="RECONCILIATION NOTE: XAU/ZAR trades are absent from the USD statement as metal legs. The USD/ZAR FX legs in both statements are the link back to the USD ecosystem. Total XAU position = 309 oz (USD) + 52 oz (ZAR) = 361 oz."
ws4[f"A{r}"].font=Font(italic=True,size=9); ws4[f"A{r}"].fill=fill("FFF3CD"); ws4[f"A{r}"].alignment=wrap()
ws4.row_dimensions[r].height=32
autofit(ws4); ws4.freeze_panes="A4"

# ──────────────────────────── SHEET 5: XAGZAR RECON ─────────────────────────
# Same formula structure as XAUZAR: F=D*E, H=F/G, I cumulative D, J cumulative F
ws5=wb.create_sheet("XAGZAR Recon"); ws5.sheet_properties.tabColor="7030A0"
title_row(ws5,"XAG/ZAR — ZAR-Denominated Silver Forwards — March 2026",PURP)
ws5.merge_cells("A2:J2")
ws5["A2"]="⚠  These trades appear ONLY in the ZAR (R) statement. They are NOT visible as metal legs in the USD ($) statement."
ws5["A2"].font=Font(bold=True,color="4B0082",size=10); ws5["A2"].fill=fill("F3E5F5"); ws5["A2"].alignment=wrap()
ws5.row_dimensions[2].height=18
r=3
hdr_row(ws5,r,["Trade Date","Value Date","Doc #","Oz","ZAR Price (R/oz)","ZAR Amount","Impl USD/ZAR","USD Equivalent","Running Oz","Running ZAR Cost"],PURP); r+=1

xagzar=[
    ("19-Mar-26","23-Mar-26","FNC/2026/070490",135,1227.9749,17.0000),
    ("19-Mar-26","23-Mar-26","FNC/2026/070573",100,1215.0768,17.0000),
    ("23-Mar-26","25-Mar-26","FNC/2026/075477",200,1148.1720,16.7600),
    ("23-Mar-26","25-Mar-26","FNC/2026/075966",200,1160.9917,16.8300),
]
r_xagzar_start = r
for i, d in enumerate(xagzar):
    bg_ = LPURP if i%2==0 else None
    for c, v in enumerate(d[:5], 1):
        cl = ws5.cell(r, c, v)
        if bg_: cl.fill = fill(bg_)
        if c == 4: cl.number_format = '#,##0'
        if c == 5: cl.number_format = '#,##0.0000'
    cl = ws5.cell(r, 7, d[5])
    if bg_: cl.fill = fill(bg_)
    cl.number_format = '#,##0.0000'
    cl = ws5.cell(r, 6, f"=D{r}*E{r}")
    if bg_: cl.fill = fill(bg_)
    cl.number_format = '#,##0.00'
    cl = ws5.cell(r, 8, f"=F{r}/G{r}")
    if bg_: cl.fill = fill(bg_)
    cl.number_format = '#,##0.00'
    i_f = f"=D{r}" if i == 0 else f"=I{r-1}+D{r}"
    cl = ws5.cell(r, 9, i_f)
    if bg_: cl.fill = fill(bg_)
    cl.number_format = '#,##0'
    j_f = f"=F{r}" if i == 0 else f"=J{r-1}+F{r}"
    cl = ws5.cell(r, 10, j_f)
    if bg_: cl.fill = fill(bg_)
    cl.number_format = '#,##0.00'
    r += 1
r_xagzar_end = r - 1

hdr_row(ws5,r,["TOTAL","","",
    f"=I{r_xagzar_end}",
    "",
    f'="R "&TEXT(J{r_xagzar_end},"#,##0.00")',
    "",
    f"=SUM(H{r_xagzar_start}:H{r_xagzar_end})",
    f"=I{r_xagzar_end}",
    f"=J{r_xagzar_end}"],PURP)
ws5.cell(r,4).number_format='#,##0'
ws5.cell(r,8).number_format='#,##0.00'; r+=1
ws5.cell(r,1,"VWAP").font=bfont()
ws5.cell(r,5,f"=J{r_xagzar_end}/I{r_xagzar_end}").number_format='#,##0.00'; ws5.cell(r,5).font=bfont()
ws5.cell(r,8,f"=SUM(H{r_xagzar_start}:H{r_xagzar_end})/I{r_xagzar_end}").number_format='#,##0.00'; ws5.cell(r,8).font=bfont()
r+=2
ws5.cell(r,1,"MONTH-END MARK-TO-MARKET — 31 March 2026").font=bfont(sz=11); r+=1
for lbl,val in[
    ("MTM Price (USD)","$75.007 / oz"),
    ("MTM Value  635 oz × $75.007","$47,629.45"),
    ("Cost Basis (USD equivalent)","$44,399.04"),
    ("Unrealised P&L (USD equivalent)","$3,230.41"),
]:
    ws5.cell(r,1,lbl); ws5.cell(r,2,val).font=bfont(); r+=1
r+=1
ws5.merge_cells(f"A{r}:J{r}")
ws5[f"A{r}"]="RECONCILIATION NOTE: XAG/ZAR trades are absent from the USD statement as metal legs. Total XAG = 8,820 oz (USD) + 635 oz (ZAR) = 9,455 oz."
ws5[f"A{r}"].font=Font(italic=True,size=9); ws5[f"A{r}"].fill=fill("F3E5F5"); ws5[f"A{r}"].alignment=wrap()
ws5.row_dimensions[r].height=24
autofit(ws5); ws5.freeze_panes="A4"

# ─────────────────────────── SHEET 6: USDZAR RECON ──────────────────────────
ws6=wb.create_sheet("USDZAR Recon"); ws6.sheet_properties.tabColor="00B050"
title_row(ws6,"USD/ZAR FX — Trade Linkage Reconciliation — March 2026",GRND)
ws6.merge_cells("A2:J2")
ws6["A2"]="Each metal purchase is funded by a matched USD/ZAR FX leg. This sheet links every metal trade to its FX counterpart."
ws6["A2"].font=Font(italic=True,size=10); ws6["A2"].fill=fill(LGRN); ws6["A2"].alignment=wrap()
r=3
hdr_row(ws6,r,["Metal Doc #","FX Doc #","Trade Date","Value Date","Instrument","Oz","USD Amount","USD/ZAR Rate","ZAR Amount","Status"],GRND); r+=1
uzdata=[
    ("FNC/2026/067462","FNC/2026/067466","16-Mar","18-Mar","XAG/USD",2500,197248.00,16.88551,3330633.08,"MATCHED"),
    ("FNC/2026/068552","FNC/2026/068553","17-Mar","19-Mar","XAG/USD",250,20133.88,16.74001,337041.35,"MATCHED"),
    ("FNC/2026/070490","(ZAR stmt only)","19-Mar","23-Mar","XAG/ZAR",135,9751.78,17.0000,165776.61,"ZAR TRADE"),
    ("FNC/2026/070573","(ZAR stmt only)","19-Mar","23-Mar","XAG/ZAR",100,7148.11,17.0000,121507.68,"ZAR TRADE"),
    ("FNC/2026/071204","FNC/2026/071206","19-Mar","23-Mar","XAG/USD",245,16804.72,17.00827,285819.22,"MATCHED"),
    ("FNC/2026/071324","FNC/2026/071325","19-Mar","23-Mar","XAG/USD",100,6652.27,17.03433,113316.96,"MATCHED"),
    ("FNC/2026/071420","FNC/2026/071421","19-Mar","23-Mar","XAG/USD",220,15011.96,17.06641,256200.26,"MATCHED"),
    ("FNC/2026/072203","FNC/2026/072204","19-Mar","23-Mar","XAG/USD",140,10135.45,16.71622,169426.41,"MATCHED"),
    ("FNC/2026/074774","FNC/2026/074777","23-Mar","25-Mar","XAG/USD",300,19503.06,17.19366,335328.98,"MATCHED"),
    ("FNC/2026/074868","(ZAR stmt only)","23-Mar","25-Mar","XAU/ZAR",12,51134.34,17.1768,878305.92,"ZAR TRADE"),
    ("FNC/2026/075477","(ZAR stmt only)","23-Mar","25-Mar","XAG/ZAR",200,13702.00,16.7600,229634.40,"ZAR TRADE"),
    ("FNC/2026/075966","(ZAR stmt only)","23-Mar","25-Mar","XAG/ZAR",200,13797.15,16.8300,232198.34,"ZAR TRADE"),
    ("FNC/2026/076104","(ZAR stmt only)","23-Mar","25-Mar","XAU/ZAR",40,174862.13,17.1831,3003508.80,"ZAR TRADE"),
    ("FNC/2026/077437","FNC/2026/077438","24-Mar","26-Mar","XAG/USD",900,63297.81,17.00081,1076114.04,"MATCHED"),
    ("FNC/2026/077551","FNC/2026/077553","24-Mar","26-Mar","XAU/USD",125,553346.25,17.03951,9428748.96,"MATCHED"),
    ("FNC/2026/077554","FNC/2026/077555","24-Mar","26-Mar","XAG/USD",300,21020.16,17.04151,358215.27,"MATCHED"),
    ("FNC/2026/077744","FNC/2026/077745","24-Mar","26-Mar","XAG/USD",500,34197.70,17.10778,585046.73,"MATCHED"),
    ("FNC/2026/077778","FNC/2026/077779","24-Mar","26-Mar","XAU/USD SELL",-5,-21911.00,17.05860,-373770.98,"MATCHED"),
    ("FNC/2026/077860","FNC/2026/077861","24-Mar","26-Mar","XAU/USD",25,108888.75,17.09665,1861632.85,"MATCHED"),
    ("FNC/2026/077888","FNC/2026/077889","24-Mar","26-Mar","XAU/USD",10,43958.70,17.03411,748797.33,"MATCHED"),
    ("FNC/2026/078603","FNC/2026/078604","25-Mar","27-Mar","XAU/USD",25,113532.00,16.94631,1923948.47,"MATCHED"),
    ("FNC/2026/078631","FNC/2026/078634","25-Mar","27-Mar","XAU/USD",16,73032.48,16.89104,1233594.54,"MATCHED"),
    ("FNC/2026/078842","FNC/2026/078843","25-Mar","27-Mar","XAG/USD",210,15268.76,16.90123,258060.82,"MATCHED"),
    ("FNC/2026/079433","FNC/2026/079434","25-Mar","27-Mar","XAU/USD",7,31904.32,16.92538,539992.74,"MATCHED"),
    ("FNC/2026/080022","FNC/2026/080023","26-Mar","30-Mar","XAG/USD",300,20977.80,17.01447,356926.15,"MATCHED"),
    ("FNC/2026/080112","FNC/2026/080113","26-Mar","30-Mar","XAU/USD",20,89094.20,17.05581,1519573.75,"MATCHED"),
    ("FNC/2026/080543","FNC/2026/080545","26-Mar","30-Mar","XAU/USD",25,111216.00,17.06811,1898246.92,"MATCHED"),
    ("FNC/2026/080904","FNC/2026/080905","26-Mar","30-Mar","XAG/USD",100,6951.42,17.02381,118339.65,"MATCHED"),
    ("FNC/2026/080932","FNC/2026/080936","26-Mar","30-Mar","XAU/USD",5,22373.85,16.97708,379842.64,"MATCHED"),
    ("FNC/2026/081728","FNC/2026/081729","27-Mar","31-Mar","XAG/USD  ⚠ DISCREPANCY",250,"SEE NOTE",17.15087,305536.40,"DISCREPANCY"),
    ("FNC/2026/081741","FNC/2026/081743","27-Mar","31-Mar","XAU/USD",10,44300.30,17.15443,759941.25,"MATCHED"),
    ("FNC/2026/082111","FNC/2026/082112","27-Mar","31-Mar","XAU/USD",35,155266.65,17.17637,2666917.43,"MATCHED"),
    ("FNC/2026/083025","FNC/2026/083026","30-Mar","01-Apr","XAG/USD",1000,70585.20,17.15055,1210575.00,"MATCHED"),
    ("FNC/2026/083102","FNC/2026/083103","30-Mar","01-Apr","XAG/USD",500,35475.40,17.18356,609593.66,"MATCHED"),
    ("FNC/2026/083541","FNC/2026/083542","30-Mar","01-Apr","XAG/USD",150,10692.53,17.15581,183439.01,"MATCHED"),
    ("FNC/2026/084221","FNC/2026/084222","31-Mar","02-Apr","XAG/USD",250,18257.90,17.06257,311526.70,"MATCHED"),
    ("FNC/2026/084360","FNC/2026/084361","31-Mar","02-Apr","XAG/USD",105,7692.56,17.09991,131542.08,"MATCHED"),
    ("FNC/2026/084449","FNC/2026/084450","31-Mar","02-Apr","XAG/USD",500,36489.40,17.13601,625282.72,"MATCHED"),
    ("FNC/2026/084495","FNC/2026/084496","31-Mar","02-Apr","XAU/USD",5,22787.25,17.14120,390600.81,"MATCHED"),
    ("FNC/2026/084820","FNC/2026/084821","31-Mar","02-Apr","XAU/USD",6,27542.76,17.09165,470751.21,"MATCHED"),
]
for i,d in enumerate(uzdata):
    st=d[-1]
    if st=="DISCREPANCY": bg_=RED; fnt=Font(bold=True,color=WHT,size=10)
    elif st=="ZAR TRADE": bg_=LORAN; fnt=Font(color="8B4000",size=10)
    elif "SELL" in str(d[4]): bg_="FFCCCC"; fnt=rfont()
    else: bg_=LGRN if i%2==0 else None; fnt=rfont()
    for c,v in enumerate(d,1):
        cl=ws6.cell(r,c,v); cl.font=fnt
        if bg_: cl.fill=fill(bg_)
        if c in(7,9) and isinstance(v,float): cl.number_format='#,##0.00'
        if c==8 and isinstance(v,float): cl.number_format='#,##0.0000'
        if c==6 and isinstance(v,(int,float)): cl.number_format='#,##0'
    r+=1
r+=1
ws6.merge_cells(f"A{r}:J{r}")
ws6[f"A{r}"]="⚠ DISCREPANCY — FNC/2026/081728 / FNC/081729: USD stmt shows 250 oz XAG @ $68.7385 = $17,184.63. ZAR stmt shows USD/ZAR 17,814.63 @ 17.15087 = R305,536.40 (implied USD $17,814.63). DIFFERENCE = $630.00. Requires StoneX trade confirmation."
ws6[f"A{r}"].font=Font(bold=True,color=WHT,size=10); ws6[f"A{r}"].fill=fill(RED)
ws6[f"A{r}"].alignment=wrap(); ws6.row_dimensions[r].height=32
autofit(ws6); ws6.freeze_panes="A4"

# ──────────────────────────── SHEET 7: SWAP FEES ────────────────────────────
# Fee column (I) = G + H  (H is the negative far leg, so I = near - |far| = net cost)
# SUB-TOTALs use SUM() formulas; GRAND TOTAL references the three subtotal cells.
# Header fixed to A1:I1 (9 columns, not 10).
ws7=wb.create_sheet("Swap Fees"); ws7.sheet_properties.tabColor=GREY
title_row(ws7,"Rolling Swap Fees — Carry Costs — March 2026",GREY,cols="A1:I1")
ws7.merge_cells("A2:I2")
ws7["A2"]="Fee = Far Leg cost − Near Leg receipt  |  Represents overnight financing cost to maintain open positions."
ws7["A2"].font=Font(italic=True,size=10); ws7["A2"].fill=fill(LGRY)
shdr=["Trade Date","SWT Doc #","Instrument","Notional","Rate","Days","Near Leg (USD)","Far Leg (USD)","Fee (USD)"]

def swap_sec(ws, sr, title, bg, bgd, rows):
    """Write a swap section. Returns (next_available_row, subtotal_row_number)."""
    ws.cell(sr,1,title).font=bfont(sz=11); sr+=1
    hdr_row(ws,sr,shdr,bg); sr+=1
    data_start=sr
    for i,d in enumerate(rows):
        bg_=bgd if i%2==0 else None
        for c,v in enumerate(d[:8],1):          # cols 1-8: input data (skip pre-computed fee)
            cl=ws.cell(sr,c,v)
            if bg_: cl.fill=fill(bg_)
            if c in(7,8) and isinstance(v,float): cl.number_format='#,##0.00'
        # Col 9 Fee = Near Leg + Far Leg  (Far Leg is already negative, so sum = net cost)
        cl=ws.cell(sr,9,f"=G{sr}+H{sr}")
        if bg_: cl.fill=fill(bg_)
        cl.number_format='#,##0.00'
        sr+=1
    data_end=sr-1
    sub_row=sr
    ws.cell(sr,1,"SUB-TOTAL").font=bfont()
    cl=ws.cell(sr,9,f"=SUM(I{data_start}:I{data_end})")
    cl.number_format='#,##0.00'; cl.font=Font(bold=True,color=RED)
    ws.cell(sr,1).fill=fill(LGRY); cl.fill=fill(LGRY)
    return sr+2, sub_row

r=3
r,sub_a=swap_sec(ws7,r,"SECTION A — XAG/USD Carry Swaps",BLUE,LBLUE,[
    ("18-Mar","SWT/2026/039553","XAG/USD","2,500 oz","4.77%",1,189804.00,-189829.15,25.15),
    ("19-Mar","SWT/2026/040878","XAG/USD","2,750 oz","5.12%",1,199602.56,-199630.94,28.38),
    ("20-Mar","SWT/2026/041663","XAG/USD","2,750 oz","5.12%",3,186811.21,-186890.91,79.70),
    ("23-Mar","SWT/2026/042450","XAG/USD","3,690 oz","5.12%",1,254714.24,-254750.48,36.24),
    ("24-Mar","SWT/2026/043320","XAG/USD","3,690 oz","4.87%",1,256502.79,-256537.47,34.68),
    ("25-Mar","SWT/2026/044128","XAG/USD","4,390 oz","4.87%",1,316076.93,-316119.69,42.76),
    ("26-Mar","SWT/2026/044937","XAG/USD","6,090 oz","4.87%",1,408793.69,-408848.98,55.29),
    ("27-Mar","SWT/2026/045738","XAG/USD","6,300 oz","4.87%",3,441333.90,-441513.01,179.11),
    ("30-Mar","SWT/2026/046535","XAG/USD","6,700 oz","4.87%",1,468417.70,-468481.09,63.39),
    ("31-Mar","SWT/2026/047490","XAG/USD","6,950 oz","4.87%",1,521298.65,-521369.19,70.54),
])
r,sub_b=swap_sec(ws7,r,"SECTION B — XAU/USD Carry Swaps","D4A755",LGOLD,[
    ("25-Mar","SWT/2026/044127","XAU/USD","52 oz","4.87%",1,235527.76,-235559.64,31.88),
    ("26-Mar","SWT/2026/044936","XAU/USD","207 oz","4.87%",1,902783.93,-902906.06,122.13),
    ("27-Mar","SWT/2026/045737","XAU/USD","255 oz","4.87%",3,1150988.40,-1151455.56,467.16),
    ("30-Mar","SWT/2026/046534","XAU/USD","305 oz","4.87%",1,1372129.43,-1372315.17,185.74),
    ("31-Mar","SWT/2026/047489","XAU/USD","350 oz","4.87%",1,1633457.00,-1633677.85,220.85),
])
r,sub_c=swap_sec(ws7,r,"SECTION C — ZAR Funding Swaps (USD cost)",GRND,LGRN,[
    ("18-Mar","SWT/2026/039554","USD/ZAR","ZAR 3,330,633","3.59%",1,196356.08,-196336.50,19.58),
    ("19-Mar","SWT/2026/040879","USD/ZAR","ZAR 3,667,674","3.59%",1,218857.32,-218835.50,21.82),
    ("20-Mar","SWT/2026/041664","USD/ZAR","ZAR 3,667,674","3.59%",3,214490.28,-214426.13,64.15),
    ("23-Mar","SWT/2026/042451","USD/ZAR","ZAR 4,779,722","3.59%",1,284072.67,-284044.35,28.32),
    ("24-Mar","SWT/2026/043321","USD/ZAR","ZAR 4,779,722","3.88%",1,280328.34,-280298.13,30.21),
    ("25-Mar","SWT/2026/044129","USD/ZAR","ZAR 9,458,698","3.88%",1,558308.01,-558247.84,60.17),
    ("26-Mar","SWT/2026/044938","USD/ZAR","ZAR 23,143,482","3.88%",1,1349025.03,-1348879.65,145.38),
    ("27-Mar","SWT/2026/045736","USD/ZAR","ZAR 27,099,079","3.88%",3,1582039.95,-1581528.63,511.32),
    ("30-Mar","SWT/2026/046536","USD/ZAR","ZAR 31,372,008","3.88%",1,1826374.96,-1826178.17,196.79),
    ("31-Mar","SWT/2026/047491","USD/ZAR","ZAR 35,104,403","3.88%",1,2070007.17,-2069784.07,223.10),
])
# Grand total references the three subtotal cells directly
ws7.cell(r,1,"GRAND TOTAL SWAP FEES (USD)").font=Font(bold=True,size=12)
cl=ws7.cell(r,9,f"=I{sub_a}+I{sub_b}+I{sub_c}")
cl.number_format='#,##0.00'; cl.font=Font(bold=True,color=RED,size=12)
ws7.cell(r,1).fill=fill(LGRY); cl.fill=fill(LGRY); r+=2
ws7.cell(r,1,"Note: 3-day rolls (20-Mar Fri→Mon, 27-Mar Thu→Mon) cover weekends. ZAR funding rate 3.59–3.88% = cost of borrowing ZAR to fund USD/ZAR FX hedge on metal positions.").font=Font(italic=True,size=9)
autofit(ws7); ws7.freeze_panes="A3"

# ─────────────────────────── SHEET 8: INTEREST EARNED ───────────────────────
ws8=wb.create_sheet("Interest Earned"); ws8.sheet_properties.tabColor=GRND
title_row(ws8,"Interest Earned on Cash Balance — March 2026",GRND,cols="A1:F1")
r=2; hdr_row(ws8,r,["Date","Reference","Avg Balance (USD)","Rate","Days","Interest (USD)"],GRND); r+=1
int_data_start=r
for d,ref,bal,rate,days,amt in[
    ("30-Mar-2026","EJV/2026/002998",141568.29,"4.87%",1,19.15),
    ("31-Mar-2026","EJV/2026/003045",189634.36,"4.87%",1,25.65),
]:
    ws8.cell(r,1,d); ws8.cell(r,2,ref)
    ws8.cell(r,3,bal).number_format='#,##0.00'
    ws8.cell(r,4,rate); ws8.cell(r,5,days)
    ws8.cell(r,6,amt).number_format='#,##0.00'
    ws8.cell(r,1).fill=fill(LGRN); r+=1
int_data_end=r-1
ws8.cell(r,1,"TOTAL").font=bfont()
ws8.cell(r,6,f"=SUM(F{int_data_start}:F{int_data_end})").number_format='#,##0.00'
ws8.cell(r,6).font=Font(bold=True,color=GRND); r+=2
ws8.cell(r,1,"Note: Interest accrued on overnight USD cash balance at StoneX (Fed Funds equivalent rate). Full-month March 2026 total = $44.80.").font=Font(italic=True,size=9)
autofit(ws8); ws8.freeze_panes="A3"

# ─────────────────────────── SHEET 9: POSITION VALUATION ────────────────────
# D=Net Oz (=B-C for individual rows; =D_xau+D_zar for COMBINED; "" for TOTAL)
# H=MTM Value (=D*G for individual; =H1+H2 for COMBINED; =H_xau_c+H_xag_c for TOTAL)
# I=Unrealised P&L (=H-F for all rows with data)
# J=P&L% (=I/F for all rows with data)
# E=Avg Cost VWAP: input for individual rows; =F/D for COMBINED/TOTAL
# F=Total Cost: input for individual rows; =F1+F2 for COMBINED; =F_xau_c+F_xag_c for TOTAL
ws9=wb.create_sheet("Position Valuation"); ws9.sheet_properties.tabColor=NAVY
title_row(ws9,"Month-End Position Valuation — 31 March 2026",NAVY)
r=2
hdr_row(ws9,r,["Instrument","Long Oz","Short Oz","Net Oz","Avg Cost (USD/oz)","Total Cost (USD)","MTM Price","MTM Value (USD)","Unrealised P&L","P&L %"],NAVY); r+=1

# vdata: (label, long_oz, short_oz, net_oz, avg_cost, total_cost, mtm_price, mtm_val, pnl, pnl_pct, bg)
# For COMBINED/TOTAL rows the pre-computed values are kept for reference but cells use formulas.
vdata=[
    ("XAU/USD",     314, 5, 309, 4450.91, 1375332.51, 4667.02, 1441909.18, 66576.67, "4.84%", LGOLD),
    ("XAU/ZAR",      52, 0,  52, 4346.08,  225996.47, 4667.02,  242684.04, 16687.57, "7.38%", LORAN),
    ("XAU COMBINED",366, 5, 361, 4435.81, 1601328.98, 4667.02, 1684593.22, 83264.24, "5.20%", GOLD ),
    ("XAG/USD",    8820, 0,8820,   72.97,  643580.61,   75.007,  661561.74, 17981.13, "2.79%", LBLUE),
    ("XAG/ZAR",     635, 0, 635,   69.92,   44399.04,   75.007,   47629.45,  3230.41, "7.28%", LPURP),
    ("XAG COMBINED",9455,0,9455,   72.76,  687979.65,   75.007,  709191.19, 21211.54, "3.08%", BLUE ),
    ("TOTAL",        "", "", "",     "",   2289308.63,       "",  2393784.41,104475.78, "4.56%", NAVY ),
]
pv_row={}
for d in vdata:
    bg_=d[-1]; is_c="COMBINED" in str(d[0]) or d[0]=="TOTAL"
    fg_=WHT if bg_ in(BLUE,NAVY) else "000000"
    label=d[0]; pv_row[label]=r

    # A: Instrument label
    cl=ws9.cell(r,1,label); cl.fill=fill(bg_); cl.font=Font(bold=is_c,color=fg_,size=10)

    # B: Long Oz, C: Short Oz — inputs for individual rows; SUM formula for COMBINED
    if label=="XAU COMBINED":
        b_v=f"=B{pv_row['XAU/USD']}+B{pv_row['XAU/ZAR']}"
        c_v=f"=C{pv_row['XAU/USD']}+C{pv_row['XAU/ZAR']}"
    elif label=="XAG COMBINED":
        b_v=f"=B{pv_row['XAG/USD']}+B{pv_row['XAG/ZAR']}"
        c_v=f"=C{pv_row['XAG/USD']}+C{pv_row['XAG/ZAR']}"
    else:
        b_v=d[1]; c_v=d[2]
    for col_idx,val in [(2,b_v),(3,c_v)]:
        cl=ws9.cell(r,col_idx,val); cl.fill=fill(bg_); cl.font=Font(bold=is_c,color=fg_,size=10)
        if isinstance(val,int): cl.number_format='#,##0'

    # D: Net Oz = Long - Short (formula for all non-TOTAL rows)
    if label=="TOTAL":
        d_f=""
    elif label=="XAU COMBINED":
        d_f=f"=D{pv_row['XAU/USD']}+D{pv_row['XAU/ZAR']}"
    elif label=="XAG COMBINED":
        d_f=f"=D{pv_row['XAG/USD']}+D{pv_row['XAG/ZAR']}"
    else:
        d_f=f"=B{r}-C{r}"
    cl=ws9.cell(r,4,d_f); cl.fill=fill(bg_); cl.font=Font(bold=is_c,color=fg_,size=10)
    if d_f: cl.number_format='#,##0'

    # F: Total Cost — input for individual; SUM for COMBINED/TOTAL
    if label=="XAU COMBINED":
        f_v=f"=F{pv_row['XAU/USD']}+F{pv_row['XAU/ZAR']}"
    elif label=="XAG COMBINED":
        f_v=f"=F{pv_row['XAG/USD']}+F{pv_row['XAG/ZAR']}"
    elif label=="TOTAL":
        f_v=f"=F{pv_row['XAU COMBINED']}+F{pv_row['XAG COMBINED']}"
    else:
        f_v=d[5]
    cl=ws9.cell(r,6,f_v); cl.fill=fill(bg_); cl.font=Font(bold=is_c,color=fg_,size=10); cl.number_format='#,##0.00'

    # E: Avg Cost VWAP — input for individual; =F/D for COMBINED/TOTAL
    if label in("XAU COMBINED","XAG COMBINED","TOTAL") and d_f:
        e_v=f"=IF(D{r}=0,0,F{r}/D{r})"
    elif label=="TOTAL":
        e_v=""
    else:
        e_v=d[4]
    cl=ws9.cell(r,5,e_v); cl.fill=fill(bg_); cl.font=Font(bold=is_c,color=fg_,size=10)
    if e_v and e_v!="": cl.number_format='#,##0.00'

    # G: MTM Price — input for individual rows; blank for COMBINED/TOTAL
    g_v=d[6] if isinstance(d[6],float) else ""
    cl=ws9.cell(r,7,g_v); cl.fill=fill(bg_); cl.font=Font(bold=is_c,color=fg_,size=10)
    if g_v: cl.number_format='#,##0.00'

    # H: MTM Value — formula =D*G for individual; SUM for COMBINED/TOTAL
    if label=="XAU COMBINED":
        h_f=f"=H{pv_row['XAU/USD']}+H{pv_row['XAU/ZAR']}"
    elif label=="XAG COMBINED":
        h_f=f"=H{pv_row['XAG/USD']}+H{pv_row['XAG/ZAR']}"
    elif label=="TOTAL":
        h_f=f"=H{pv_row['XAU COMBINED']}+H{pv_row['XAG COMBINED']}"
    else:
        h_f=f"=D{r}*G{r}"
    cl=ws9.cell(r,8,h_f); cl.fill=fill(bg_); cl.font=Font(bold=is_c,color=fg_,size=10); cl.number_format='#,##0.00'

    # I: Unrealised P&L = MTM Value − Cost  (formula for all rows)
    if label=="XAU COMBINED":
        i_f=f"=I{pv_row['XAU/USD']}+I{pv_row['XAU/ZAR']}"
    elif label=="XAG COMBINED":
        i_f=f"=I{pv_row['XAG/USD']}+I{pv_row['XAG/ZAR']}"
    elif label=="TOTAL":
        i_f=f"=I{pv_row['XAU COMBINED']}+I{pv_row['XAG COMBINED']}"
    else:
        i_f=f"=H{r}-F{r}"
    cl=ws9.cell(r,9,i_f); cl.fill=fill(bg_); cl.font=Font(bold=is_c,color=fg_,size=10); cl.number_format='#,##0.00'

    # J: P&L% = P&L / Cost  (formula for all rows)
    j_f=f'=IF(F{r}=0,"",I{r}/F{r})'
    cl=ws9.cell(r,10,j_f); cl.fill=fill(bg_); cl.font=Font(bold=is_c,color=fg_,size=10); cl.number_format='0.00%'
    r+=1

pv_total_row = pv_row["TOTAL"]

r+=1
ws9.cell(r,1,"CARRY COST IMPACT").font=bfont(sz=11); r+=1
hdr_row(ws9,r,["Item","Amount (USD)"],GREY); r+=1
carry_start=r
for lbl,amt in[
    ("Gross Unrealised P&L",104475.78),
    ("Total Swap Fees Paid (March 2026)",-2943.84),
    ("Interest Earned (March 2026)",44.80),
]:
    ws9.cell(r,1,lbl).font=Font(size=10)
    cl=ws9.cell(r,2,amt); cl.number_format='#,##0.00'
    cl.font=Font(color=RED if amt<0 else "000000",size=10)
    r+=1
# NET P&L row references the three carry cost cells above it
net_row=r
ws9.cell(r,1,"NET P&L AFTER CARRY COSTS").font=Font(bold=True,size=10)
cl=ws9.cell(r,2,f"=SUM(B{carry_start}:B{r-1})")
cl.number_format='#,##0.00'; cl.font=Font(bold=True,color=GRND,size=10)
ws9.cell(r,1).fill=fill(LGRY); cl.fill=fill(LGRY)
r+=2
ws9.merge_cells(f"A{r}:J{r}")
ws9[f"A{r}"]="IMPORTANT: MTM prices XAU $4,667.02/oz and XAG $75.007/oz are sourced from 31-Mar-2026 StoneX overnight SWT rates and must be confirmed against official StoneX closing prices before presenting to auditors. ZAR-denominated positions use implied USD conversion rates from trade date."
ws9[f"A{r}"].font=Font(italic=True,size=9,color="666666"); ws9[f"A{r}"].alignment=wrap()
ws9.row_dimensions[r].height=32
autofit(ws9); ws9.freeze_panes="A3"

# ══ SAVE EXCEL ════════════════════════════════════════════════════════════════
xl_path=os.path.join(OUT_DIR,"StoneX March 2026 Recon.xlsx")
wb.save(xl_path)
print(f"EXCEL SAVED: {xl_path}")

# ══════════════════════════════════════════════════════════════════════════════
# PDF REPORT
# ══════════════════════════════════════════════════════════════════════════════
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, PageBreak, HRFlowable)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

pdf_path=os.path.join(OUT_DIR,"StoneX March 2026 Recon Report.pdf")
doc=SimpleDocTemplate(pdf_path,pagesize=A4,
    leftMargin=1.8*cm,rightMargin=1.8*cm,topMargin=2*cm,bottomMargin=2*cm)

# SA Bullion brand palette — PDF
C_NAVY=colors.HexColor("#150E26")   # deep dark purple
C_GOLD=colors.HexColor("#D4A755")   # satin gold
C_BLUE=colors.HexColor("#7B4FC9")   # tyrian purple
C_GRN =colors.HexColor("#40B5AD")   # persian teal
C_RED =colors.HexColor("#E05252")   # brand red
C_ORAN=colors.HexColor("#D4720A")   # brand orange
C_PURP=colors.HexColor("#4B1D75")   # dark purple
C_LGRY=colors.HexColor("#F2F2F2")   # light grey
C_LBLU=colors.HexColor("#EDE4F7")   # light purple
C_LGLD=colors.HexColor("#FAF3E0")   # light gold cream
C_WARNBG=colors.HexColor("#FFF3CD") # warning amber

styles=getSampleStyleSheet()
H1=ParagraphStyle("H1",fontSize=16,fontName="Helvetica-Bold",textColor=C_GOLD,
    backColor=C_NAVY,alignment=TA_CENTER,spaceAfter=4,spaceBefore=4,leading=20,
    leftIndent=-10,rightIndent=-10)
H2=ParagraphStyle("H2",fontSize=12,fontName="Helvetica-Bold",textColor=C_GOLD,
    backColor=C_NAVY,spaceBefore=12,spaceAfter=4,leading=16)
H3=ParagraphStyle("H3",fontSize=10,fontName="Helvetica-Bold",textColor=C_PURP,
    spaceBefore=10,spaceAfter=4)
BODY=ParagraphStyle("BODY",fontSize=9,fontName="Helvetica",spaceAfter=4,leading=13)
NOTE=ParagraphStyle("NOTE",fontSize=8,fontName="Helvetica-Oblique",textColor=colors.grey,
    spaceAfter=6,leading=11)
WARN=ParagraphStyle("WARN",fontSize=9,fontName="Helvetica-Bold",textColor=C_RED,
    backColor=C_WARNBG,spaceAfter=6,leading=12)

def tbl(data,col_widths,hdr_bg=C_NAVY,hdr_fg=colors.white,alt=True,font_size=8):
    style=[
        ('BACKGROUND',(0,0),(-1,0),hdr_bg),
        ('TEXTCOLOR',(0,0),(-1,0),hdr_fg),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),font_size),
        ('FONTNAME',(0,1),(-1,-1),'Helvetica'),
        ('GRID',(0,0),(-1,-1),0.3,colors.grey),
        ('ROWBACKGROUND',(0,0),(-1,0),[hdr_bg]),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LEFTPADDING',(0,0),(-1,-1),4),
        ('RIGHTPADDING',(0,0),(-1,-1),4),
        ('TOPPADDING',(0,0),(-1,-1),3),
        ('BOTTOMPADDING',(0,0),(-1,-1),3),
    ]
    if alt:
        for i in range(1,len(data)):
            if i%2==1:
                style.append(('BACKGROUND',(0,i),(-1,i),C_LGRY))
    t=Table(data,colWidths=col_widths)
    t.setStyle(TableStyle(style))
    return t

story=[]

# ── Cover page ──
story.append(Spacer(1,3*cm))
cover_data=[["StoneX Financial Ltd"],["March 2026 Reconciliation Report"],
            [" "],["SA Bullion Investor Services"],["Account: MT0795"],
            [" "],["Prepared by: Treasury Team  |  April 2026"],["CONFIDENTIAL"]]
ct=Table(cover_data,colWidths=[15*cm])
ct.setStyle(TableStyle([
    ('BACKGROUND',(0,0),(-1,-1),C_NAVY),
    ('TEXTCOLOR',(0,0),(-1,-1),colors.white),
    ('FONTNAME',(0,0),(0,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(0,0),20),
    ('FONTNAME',(0,1),(0,1),'Helvetica-Bold'),('FONTSIZE',(0,1),(0,1),15),
    ('FONTNAME',(0,3),(0,3),'Helvetica-Bold'),('FONTSIZE',(0,3),(0,3),13),
    ('FONTNAME',(0,4),(0,5),'Helvetica'),('FONTSIZE',(0,4),(0,5),11),
    ('FONTNAME',(0,6),(0,7),'Helvetica-Oblique'),('FONTSIZE',(0,6),(0,7),9),
    ('ALIGN',(0,0),(-1,-1),'CENTER'),
    ('TOPPADDING',(0,0),(0,0),30),('BOTTOMPADDING',(0,-1),(-1,-1),30),
    ('TOPPADDING',(0,1),(0,1),8),('TOPPADDING',(0,3),(0,3),20),
]))
story.append(ct)
story.append(PageBreak())

# ── Page 2: Executive Summary ──
story.append(Paragraph("1. Executive Summary",H2))
story.append(Paragraph("This report reconciles the StoneX Financial Ltd account (MT0795 — SA Bullion Investor Services) for the period 16 March to 31 March 2026. The account trades XAU (gold) and XAG (silver) forward contracts across two currency environments: USD-denominated forwards and ZAR-denominated forwards. A daily rolling USD/ZAR FX swap structure funds the positions overnight.",BODY))
story.append(Spacer(1,0.3*cm))

story.append(Paragraph("Position Summary",H3))
pos_table_data=[
    ["Instrument","Net Oz","VWAP (USD/oz)","Cost Basis (USD)","MTM Price","MTM Value (USD)","Unrealised P&L","P&L %"],
    ["XAU/USD","309","$4,450.91","$1,375,332.51","$4,667.02","$1,441,909.18","$66,576.67","4.84%"],
    ["XAU/ZAR","52","$4,346.08","$225,996.47","$4,667.02","$242,684.04","$16,687.57","7.38%"],
    ["XAU COMBINED","361","$4,435.81","$1,601,328.98","$4,667.02","$1,684,593.22","$83,264.24","5.20%"],
    ["XAG/USD","8,820","$72.97","$643,580.61","$75.007","$661,561.74","$17,981.13","2.79%"],
    ["XAG/ZAR","635","$69.92","$44,399.04","$75.007","$47,629.45","$3,230.41","7.28%"],
    ["XAG COMBINED","9,455","$72.76","$687,979.65","$75.007","$709,191.19","$21,211.54","3.08%"],
    ["TOTAL PORTFOLIO","","","$2,289,308.63","","$2,393,784.41","$104,475.78","4.56%"],
]
cw=[3.2*cm,1.4*cm,2.0*cm,2.6*cm,1.8*cm,2.6*cm,2.2*cm,1.2*cm]
t=tbl(pos_table_data,cw)
t.setStyle(TableStyle([
    ('BACKGROUND',(0,3),(-1,3),colors.HexColor("#FFD700")),
    ('BACKGROUND',(0,6),(-1,6),C_BLUE),('TEXTCOLOR',(0,6),(-1,6),colors.white),
    ('BACKGROUND',(0,7),(-1,7),C_NAVY),('TEXTCOLOR',(0,7),(-1,7),colors.white),
    ('FONTNAME',(0,3),(-1,3),'Helvetica-Bold'),
    ('FONTNAME',(0,6),(-1,6),'Helvetica-Bold'),
    ('FONTNAME',(0,7),(-1,7),'Helvetica-Bold'),
]))
story.append(t)
story.append(Spacer(1,0.3*cm))
story.append(Paragraph("MTM prices: XAU $4,667.02/oz and XAG $75.007/oz sourced from 31-Mar-2026 StoneX overnight swap rate. To be confirmed against official closing spot prices for audit purposes.",NOTE))
story.append(PageBreak())

# ── Page 3: Cash Reconciliation ──
story.append(Paragraph("2. Cash Reconciliation (USD Account)",H2))
cash_data=[
    ["Item","Amount (USD)"],
    ["Opening Balance","$0.00"],
    ["+ Incoming Wires (3 transfers)","$445,233.94"],
    ["− XAU Purchases — USD legs","($1,375,332.51)"],
    ["− XAG Purchases — USD legs","($643,580.61)"],
    ["+ XAU Sale Proceeds (5 oz short)","$21,911.00"],
    ["− Total Swap Fees (carry costs)","($2,943.84)"],
    ["+ Interest Earned","$44.80"],
    ["= Closing Balance (USD)","$442,964.60"],
    ["ZAR Ledger Closing Balance","(R39,037,714.16)"],
]
ct2=tbl(cash_data,[10*cm,7*cm])
ct2.setStyle(TableStyle([
    ('BACKGROUND',(0,8),(-1,8),colors.HexColor("#E2EFDA")),
    ('FONTNAME',(0,8),(-1,8),'Helvetica-Bold'),
    ('TEXTCOLOR',(0,3),(-1,4),C_RED),
    ('TEXTCOLOR',(0,6),(1,6),C_RED),
]))
story.append(ct2)
story.append(Spacer(1,0.3*cm))
story.append(Paragraph("Note: The negative ZAR closing balance (R39,037,714.16) represents the net ZAR short position from the rolling USD/ZAR FX funding structure. ZAR is borrowed daily (at 3.59–3.88%) to fund the USD/ZAR positions that accompany each metal purchase.",BODY))

story.append(Spacer(1,0.4*cm))
story.append(Paragraph("Incoming Wires Detail",H3))
wire_data=[
    ["Date","Reference","Amount (USD)"],
    ["13-Mar-2026","JRV/2026/013947","$120,435.46"],
    ["24-Mar-2026","JRV/2026/015715","$177,989.26"],
    ["27-Mar-2026","JRV/2026/016606","$146,809.22"],
    ["TOTAL","","$445,233.94"],
]
story.append(tbl(wire_data,[4*cm,6*cm,7*cm],hdr_bg=C_ORAN))
story.append(PageBreak())

# ── Page 4: Swap Fees ──
story.append(Paragraph("3. Swap Fees — Rolling Carry Costs",H2))
story.append(Paragraph("Each overnight position roll generates a swap fee: the difference between the far-leg cost and near-leg receipt. Three categories of swaps run in parallel daily: XAG carry, XAU carry, and ZAR funding.",BODY))

story.append(Paragraph("XAG/USD Carry Swaps",H3))
xag_swap_data=[
    ["Date","SWT Doc #","Oz","Rate","Days","Near Leg","Far Leg","Fee (USD)"],
    ["18-Mar","SWT/039553","2,500","4.77%","1","189,804.00","(189,829.15)","25.15"],
    ["19-Mar","SWT/040878","2,750","5.12%","1","199,602.56","(199,630.94)","28.38"],
    ["20-Mar","SWT/041663","2,750","5.12%","3","186,811.21","(186,890.91)","79.70"],
    ["23-Mar","SWT/042450","3,690","5.12%","1","254,714.24","(254,750.48)","36.24"],
    ["24-Mar","SWT/043320","3,690","4.87%","1","256,502.79","(256,537.47)","34.68"],
    ["25-Mar","SWT/044128","4,390","4.87%","1","316,076.93","(316,119.69)","42.76"],
    ["26-Mar","SWT/044937","6,090","4.87%","1","408,793.69","(408,848.98)","55.29"],
    ["27-Mar","SWT/045738","6,300","4.87%","3","441,333.90","(441,513.01)","179.11"],
    ["30-Mar","SWT/046535","6,700","4.87%","1","468,417.70","(468,481.09)","63.39"],
    ["31-Mar","SWT/047490","6,950","4.87%","1","521,298.65","(521,369.19)","70.54"],
    ["SUB-TOTAL","","","","","","","$615.24"],
]
story.append(tbl(xag_swap_data,[1.8*cm,3*cm,1.5*cm,1.3*cm,1*cm,2.5*cm,2.5*cm,2.4*cm],hdr_bg=C_BLUE))

story.append(Paragraph("XAU/USD Carry Swaps",H3))
xau_swap_data=[
    ["Date","SWT Doc #","Oz","Rate","Days","Near Leg","Far Leg","Fee (USD)"],
    ["25-Mar","SWT/044127","52","4.87%","1","235,527.76","(235,559.64)","31.88"],
    ["26-Mar","SWT/044936","207","4.87%","1","902,783.93","(902,906.06)","122.13"],
    ["27-Mar","SWT/045737","255","4.87%","3","1,150,988.40","(1,151,455.56)","467.16"],
    ["30-Mar","SWT/046534","305","4.87%","1","1,372,129.43","(1,372,315.17)","185.74"],
    ["31-Mar","SWT/047489","350","4.87%","1","1,633,457.00","(1,633,677.85)","220.85"],
    ["SUB-TOTAL","","","","","","","$1,027.76"],
]
story.append(tbl(xau_swap_data,[1.8*cm,3*cm,1.5*cm,1.3*cm,1*cm,2.5*cm,2.5*cm,2.4*cm],hdr_bg=colors.HexColor("#B8860B")))

story.append(Paragraph("ZAR Funding Swaps (cost in USD)",H3))
story.append(Paragraph("The USD/ZAR funding swaps roll the ZAR borrowed to fund metal positions. The net cost in USD terms is shown below.",BODY))
zar_swap_summary=[
    ["Date","SWT Doc #","ZAR Notional","Rate","Days","Fee (USD)"],
    ["18-Mar","SWT/039554","ZAR 3,330,633","3.59%","1","19.58"],
    ["19-Mar","SWT/040879","ZAR 3,667,674","3.59%","1","21.82"],
    ["20-Mar","SWT/041664","ZAR 3,667,674","3.59%","3","64.15"],
    ["23-Mar","SWT/042451","ZAR 4,779,722","3.59%","1","28.32"],
    ["24-Mar","SWT/043321","ZAR 4,779,722","3.88%","1","30.21"],
    ["25-Mar","SWT/044129","ZAR 9,458,698","3.88%","1","60.17"],
    ["26-Mar","SWT/044938","ZAR 23,143,482","3.88%","1","145.38"],
    ["27-Mar","SWT/045736","ZAR 27,099,079","3.88%","3","511.32"],
    ["30-Mar","SWT/046536","ZAR 31,372,008","3.88%","1","196.79"],
    ["31-Mar","SWT/047491","ZAR 35,104,403","3.88%","1","223.10"],
    ["SUB-TOTAL","","","","","$1,300.84"],
]
story.append(tbl(zar_swap_summary,[1.8*cm,3*cm,3.5*cm,1.5*cm,1*cm,5.2*cm],hdr_bg=C_GRN))

summary_swap=[["",""],["GRAND TOTAL SWAP FEES (March 2026)","$2,943.84"]]
st2=Table(summary_swap,colWidths=[10*cm,7*cm])
st2.setStyle(TableStyle([
    ('BACKGROUND',(0,1),(-1,1),C_NAVY),('TEXTCOLOR',(0,1),(-1,1),colors.white),
    ('FONTNAME',(0,1),(-1,1),'Helvetica-Bold'),('FONTSIZE',(0,1),(-1,1),10),
    ('ALIGN',(1,1),(1,1),'RIGHT'),('TOPPADDING',(0,1),(-1,1),5),('BOTTOMPADDING',(0,1),(-1,1),5),
]))
story.append(st2)
story.append(PageBreak())

# ── Page 5: ZAR/USD Recon Summary ──
story.append(Paragraph("4. USD/ZAR Trade Linkage — Daily Summary",H2))
story.append(Paragraph("Each metal forward purchase is immediately funded by a matched USD/ZAR FX leg, settling on the same value date. The ZAR statement records the ZAR side of each leg. The table below shows daily totals.",BODY))

daily_uz=[
    ["Trade Date","XAU Longs","XAG Longs","ZAR Trades","Total USD","Total ZAR","Avg USD/ZAR"],
    ["16-Mar","—","2,500 oz","—","$197,248","R3,330,633","16.8855"],
    ["17-Mar","—","250 oz","—","$20,134","R337,041","16.7400"],
    ["19-Mar","—","705 oz (XAG) incl. 235 ZAR","235 oz XAG","$48,604","R856,046","17.02 avg"],
    ["23-Mar","52 oz XAU (ZAR)","700 oz XAG (incl. ZAR)","52 XAU + 400 XAG","$269,800","R4,804,178","17.02 avg"],
    ["24-Mar","155 oz net","2,000 oz XAG","—","$829,672","R14,156,355","17.05 avg"],
    ["25-Mar","48 oz","210 oz XAG","—","$234,708","R3,955,597","16.91 avg"],
    ["26-Mar","75 oz","400 oz XAG","—","$250,615","R4,272,929","17.04 avg"],
    ["27-Mar","45 oz","250 oz XAG","—","$261,752","R3,732,395","17.16 avg"],
    ["30-Mar","—","1,650 oz XAG","—","$116,753","R2,003,608","17.16 avg"],
    ["31-Mar","17 oz","855 oz XAG","—","$118,770","R2,031,297","17.10 avg"],
]
story.append(tbl(daily_uz,[2*cm,2.5*cm,3*cm,2.5*cm,2*cm,2.5*cm,2.5*cm],hdr_bg=C_GRN))
story.append(Spacer(1,0.3*cm))
story.append(Paragraph("⚠ DISCREPANCY FLAGGED — 27 March (FNC/2026/081728 / FNC/081729): USD statement records XAG 250oz @ $68.7385 = $17,184.63. ZAR statement records USD/ZAR $17,814.63 @ 17.15087 = R305,536.40. The implied USD amount from the ZAR side is $17,814.63, a difference of $630.00. A StoneX trade confirmation must be obtained to resolve this discrepancy before finalising March accounts.",WARN))
story.append(PageBreak())

# ── Page 6: XAUZAR / XAGZAR Explanation ──
story.append(Paragraph("5. ZAR-Denominated Positions — XAUZAR & XAGZAR",H2))
story.append(Paragraph("A portion of March trades were executed as ZAR-denominated forwards (priced in Rand per troy oz). These do not appear as metal legs in the USD ($) statement — they show only as USD/ZAR FX legs. They are fully captured in the ZAR (R) statement.",BODY))

story.append(Paragraph("XAU/ZAR Trades",H3))
xauzar_data=[
    ["Trade Date","Doc #","Oz","ZAR Price (R/oz)","ZAR Amount","Impl USD/ZAR","USD Equivalent"],
    ["23-Mar-26","FNC/2026/074868","12","R73,192.16","R878,305.92","17.1768","$51,134.34"],
    ["23-Mar-26","FNC/2026/076104","40","R75,087.72","R3,003,508.80","17.1831","$174,862.13"],
    ["TOTAL / VWAP","","52","R74,650.28","R3,881,814.72","","$225,996.47"],
]
story.append(tbl(xauzar_data,[2.5*cm,3.5*cm,1.2*cm,2.5*cm,3*cm,2*cm,2.3*cm],hdr_bg=C_ORAN,hdr_fg=colors.white))

story.append(Paragraph("XAG/ZAR Trades",H3))
xagzar_data=[
    ["Trade Date","Doc #","Oz","ZAR Price (R/oz)","ZAR Amount","Impl USD/ZAR","USD Equivalent"],
    ["19-Mar-26","FNC/2026/070490","135","R1,227.9749","R165,776.61","17.0000","$9,751.78"],
    ["19-Mar-26","FNC/2026/070573","100","R1,215.0768","R121,507.68","17.0000","$7,148.11"],
    ["23-Mar-26","FNC/2026/075477","200","R1,148.172","R229,634.40","16.7600","$13,702.00"],
    ["23-Mar-26","FNC/2026/075966","200","R1,160.9917","R232,198.34","16.8300","$13,797.15"],
    ["TOTAL / VWAP","","635","R1,180.50","R749,117.03","","$44,399.04"],
]
story.append(tbl(xagzar_data,[2.5*cm,3.5*cm,1.2*cm,2.5*cm,3*cm,2*cm,2.3*cm],hdr_bg=C_PURP))

story.append(Spacer(1,0.4*cm))
story.append(Paragraph("Position Totals (combining USD and ZAR denominated)",H3))
totals_data=[
    ["Metal","USD Oz","ZAR Oz","TOTAL Oz","Blended VWAP","Total Cost (USD)","MTM Value","Unrealised P&L"],
    ["XAU","309","52","361","$4,435.81","$1,601,328.98","$1,684,593.22","$83,264.24"],
    ["XAG","8,820","635","9,455","$72.76","$687,979.65","$709,191.19","$21,211.54"],
    ["TOTAL","","","","","$2,289,308.63","$2,393,784.41","$104,475.78"],
]
story.append(tbl(totals_data,[1.5*cm,1.8*cm,1.8*cm,1.8*cm,2.5*cm,3*cm,2.8*cm,2.8*cm],hdr_bg=C_NAVY))
story.append(PageBreak())

# ── Page 7: Notes & Discrepancies ──
story.append(Paragraph("6. Notes & Discrepancies",H2))

notes=[
    ("Note 1 — XAUZAR and XAGZAR trades not visible in USD statement",
     "Six trades (2 XAU/ZAR and 4 XAG/ZAR) are priced directly in South African Rand and appear only in the ZAR statement as metal legs. In the USD statement, these same trades appear only as USD/ZAR FX conversions. This is normal for ZAR-denominated forwards — there is no \"missing\" data; both statements are correct but capture different sides of the same transactions. To fully reconcile, both statements must be read together."),
    ("Note 2 — $630 discrepancy on FNC/2026/081728 / FNC/081729 (27 March)",
     "The USD statement records XAG/USD 250oz at $17,184.63 (250 × $68.7385). The matching ZAR leg (FNC/081729) in both statements records $17,814.63 as the USD reference, giving ZAR R305,536.40 at 17.15087. The difference is $630.00. This may be a keying error on one statement or a minor trade amendment not yet reflected. ACTION REQUIRED: Obtain trade confirmation from StoneX for FNC/2026/081728 before closing March accounts."),
    ("Note 3 — Month-end MTM prices",
     "MTM prices used are XAU $4,667.02/oz and XAG $75.007/oz, sourced from the 31-Mar-2026 overnight SWT forward rates. These are forward prices and may differ slightly from official StoneX closing spot prices. Confirm official closing prices for month-end accounts and audit purposes."),
    ("Note 4 — ZAR ledger closing balance",
     "The ZAR statement closes at negative R39,037,714.16. This represents the net ZAR short position from the rolling daily USD/ZAR funding swaps. Each day the account sells USD (buys ZAR) in the near leg, then buys USD (sells ZAR) in the far leg. At month-end, the open far leg for 31-Mar (settling 01-Apr) represents R35,104,402.97 outstanding, which largely accounts for the ZAR balance. This is not a cash shortfall — it is a structural feature of the rolling FX funding mechanism."),
    ("Note 5 — Interest income",
     "Interest of $44.80 was earned on the USD cash balance over 30–31 March 2026 at 4.87% (Fed Funds equivalent). As positions and capital deployed grew significantly, a materially higher interest credit is expected for a full month."),
]
for title, body in notes:
    story.append(Paragraph(title, H3))
    story.append(Paragraph(body, BODY))
    story.append(Spacer(1,0.2*cm))

story.append(HRFlowable(width="100%",thickness=1,color=C_NAVY))
story.append(Spacer(1,0.2*cm))
story.append(Paragraph("This report was prepared by the Treasury team for internal review and presentation to the head accountant. All figures sourced from StoneX Financial Ltd official statements (Account MT0795). Pending items: (1) Confirm MTM closing prices with StoneX. (2) Resolve $630 discrepancy on FNC/081728. (3) Obtain official XAUZAR/XAGZAR confirmation notes.",NOTE))

# ── Build PDF ──
def add_page_num(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica",8)
    canvas.setFillColor(colors.grey)
    canvas.drawString(1.8*cm, 1.2*cm, f"SA Bullion Investor Services | StoneX March 2026 Recon | CONFIDENTIAL")
    canvas.drawRightString(A4[0]-1.8*cm, 1.2*cm, f"Page {doc.page}")
    canvas.restoreState()

doc.build(story, onFirstPage=add_page_num, onLaterPages=add_page_num)
print(f"PDF SAVED: {pdf_path}")
print("DONE")

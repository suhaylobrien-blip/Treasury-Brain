"""
Treasury Brain — Excel Report Writer
Auto-populates the designated Excel report files per entity+metal.
"""

import json
import os
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

CONFIG_PATH = os.path.join(os.path.dirname(__file__), '..', 'config', 'settings.json')
with open(CONFIG_PATH) as f:
    CONFIG = json.load(f)

EXCEL_PATHS = CONFIG.get('excel_reports', {})
REPORTS_DIR = os.path.join(os.path.dirname(__file__), '..', CONFIG.get('reports_folder', 'reports'))


def _get_report_path(entity: str, metal: str) -> str:
    key = f"{entity}_{metal}"
    path = EXCEL_PATHS.get(key)
    if not path:
        raise ValueError(f"No Excel report path configured for {key}. "
                         f"Add it to config/settings.json → excel_reports.")
    return path


def write_daily_summary_excel(entity: str, metal: str, summary: dict):
    """
    Writes today's daily summary into the configured Excel report file.
    Creates a new row per day in the 'Daily Summary' sheet.
    """
    try:
        path = _get_report_path(entity, metal)
        if not os.path.exists(path):
            print(f"[excel_writer] Report file not found: {path}")
            return

        wb = openpyxl.load_workbook(path)

        if 'Daily Summary' not in wb.sheetnames:
            ws = wb.create_sheet('Daily Summary')
            _write_summary_headers(ws)
        else:
            ws = wb['Daily Summary']

        # Find next empty row
        row = ws.max_row + 1

        ws.cell(row, 1,  summary['date'])
        ws.cell(row, 2,  summary['deal_count'])
        ws.cell(row, 3,  summary['buy_oz'])
        ws.cell(row, 4,  summary['sell_oz'])
        ws.cell(row, 5,  summary['buy_value_zar'])
        ws.cell(row, 6,  summary['sell_value_zar'])
        ws.cell(row, 7,  summary['buy_vwap'])
        ws.cell(row, 8,  summary['sell_vwap'])
        ws.cell(row, 9,  summary['buy_margin_vwap'])
        ws.cell(row, 10, summary['sell_margin_vwap'])
        ws.cell(row, 11, summary['total_gp_zar'])
        ws.cell(row, 12, summary['inventory_oz'])
        ws.cell(row, 13, summary['spot_price_zar'])
        ws.cell(row, 14, summary['inventory_value_zar'])
        ws.cell(row, 15, summary['provision_mode']['mode'])
        ws.cell(row, 16, summary['provision_mode']['rate_pct'])

        wb.save(path)
        print(f"[excel_writer] Updated: {path}")

    except Exception as e:
        print(f"[excel_writer] Write failed for {entity} {metal}: {e}")


def _write_summary_headers(ws):
    headers = [
        'Date', 'Deals', 'Buy Oz', 'Sell Oz',
        'Buy Value (ZAR)', 'Sell Value (ZAR)',
        'Buy VWAP', 'Sell VWAP',
        'Buy Margin VWAP', 'Sell Margin VWAP',
        'Total GP (ZAR)', 'Inventory Oz',
        'Spot (ZAR)', 'Inventory Value (ZAR)',
        'Provision Mode', 'Provision Rate %',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='1F4E79')
        cell.font = Font(bold=True, color='FFFFFF')


def generate_daily_report(summaries: list) -> str:
    """
    Generates a consolidated daily Excel report across all entities and metals.
    Saves to reports/ folder. Returns the file path.
    """
    os.makedirs(REPORTS_DIR, exist_ok=True)
    today = date.today().isoformat()
    filename = f"Treasury_Daily_Report_{today}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Daily Report'

    # Header
    ws.merge_cells('A1:P1')
    title_cell = ws['A1']
    title_cell.value = f"Treasury Brain — Daily Report — {today}"
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='1F4E79')
    title_cell.alignment = Alignment(horizontal='center')

    _write_summary_headers(ws)
    # Shift headers to row 2 (title takes row 1)
    ws.delete_rows(2)
    _write_summary_headers(ws)  # re-write at row 2

    for row_num, s in enumerate(summaries, 3):
        ws.cell(row_num, 1,  s['date'])
        ws.cell(row_num, 2,  f"{s.get('entity','')} {s.get('metal','')}")
        ws.cell(row_num, 3,  s['buy_oz'])
        ws.cell(row_num, 4,  s['sell_oz'])
        ws.cell(row_num, 5,  s['buy_value_zar'])
        ws.cell(row_num, 6,  s['sell_value_zar'])
        ws.cell(row_num, 7,  s['buy_vwap'])
        ws.cell(row_num, 8,  s['sell_vwap'])
        ws.cell(row_num, 9,  s['buy_margin_vwap'])
        ws.cell(row_num, 10, s['sell_margin_vwap'])
        ws.cell(row_num, 11, s['total_gp_zar'])
        ws.cell(row_num, 12, s['inventory_oz'])
        ws.cell(row_num, 13, s['spot_price_zar'])
        ws.cell(row_num, 14, s['inventory_value_zar'])
        ws.cell(row_num, 15, s['provision_mode']['mode'])
        ws.cell(row_num, 16, s['provision_mode']['rate_pct'])

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

    wb.save(filepath)
    print(f"[excel_writer] Daily report saved: {filepath}")
    return filepath


def write_deals_excel(entity: str, metal: str, deals: list):
    """
    Appends deal rows to the 'Deals' sheet in the configured Excel report.
    """
    if not deals:
        return
    try:
        path = _get_report_path(entity, metal)
        if not os.path.exists(path):
            print(f"[excel_writer] Report file not found: {path}")
            return

        wb = openpyxl.load_workbook(path)
        if 'Deals' not in wb.sheetnames:
            ws = wb.create_sheet('Deals')
            headers = [
                'deal_id', 'deal_date', 'deal_type', 'entity', 'metal',
                'dealer_name', 'silo', 'channel', 'product_code', 'product_name',
                'units', 'equiv_oz_per_unit', 'oz', 'spot_price_zar',
                'margin_pct', 'effective_price_zar', 'deal_value_zar',
                'provision_pct', 'profit_margin_pct', 'gp_contribution_zar',
                'inventory_after_oz', 'provision_flipped',
            ]
            for col, h in enumerate(headers, 1):
                ws.cell(1, col, h).font = Font(bold=True)
        else:
            ws = wb['Deals']
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        for deal in deals:
            row = ws.max_row + 1
            for col, h in enumerate(headers, 1):
                ws.cell(row, col, deal.get(h, ''))

        wb.save(path)
        print(f"[excel_writer] Appended {len(deals)} deals to {path}")

    except Exception as e:
        print(f"[excel_writer] Deal write failed for {entity} {metal}: {e}")

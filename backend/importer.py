"""
Treasury Brain — Excel Importer
Reads SABIS dealer dealing sheets, classifies rows by colour, and writes
confirmed deals to the deals table and quote (pipeline) rows to pipeline.

Sheet layouts (confirmed with actual sheet inspection):
  KR Dealing      — buybacks cols A–N (0–13),  sales cols O–AA (14–26)
  Silver Bullion  — buybacks cols A–P (0–15),  sales cols Q–AE (16–30)
  Gold Excl KRs   — buybacks cols A–P (0–15),  sales cols Q–AE (16–30)

Row colour rules:
  Orange (R>180, G<200, B<80) → confirmed deal → deals table
  Goldish-yellow (R>200, G≥200, B<130) → quote → pipeline table
  Blue (B>130, B>R+40)        → proof coin  → deals table (product_type=proof)
  White / uncoloured          → skip (header or blank)
"""

import os
import shutil
import json
import tempfile
from datetime import datetime, date

import pandas as pd
from openpyxl import load_workbook

from processor import get_provision_mode, calc_gp_contribution, margin_vs_provision
from models import (
    get_inventory, get_latest_spot, insert_deal, insert_pipeline,
    update_inventory, insert_cash_flow,
)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

_DIR = os.path.dirname(__file__)
CONFIG_PATH   = os.path.join(_DIR, '..', 'config', 'settings.json')
PRODUCTS_PATH = os.path.join(_DIR, '..', 'config', 'products.json')

with open(CONFIG_PATH) as f:
    CONFIG = json.load(f)
with open(PRODUCTS_PATH) as f:
    _PRODUCTS = json.load(f)

INBOX     = os.path.abspath(os.path.join(_DIR, '..', CONFIG['inbox_folder']))
PROCESSED = os.path.abspath(os.path.join(_DIR, '..', CONFIG['processed_folder']))
ERRORS    = os.path.abspath(os.path.join(_DIR, '..', CONFIG['errors_folder']))

os.makedirs(INBOX,     exist_ok=True)
os.makedirs(PROCESSED, exist_ok=True)
os.makedirs(ERRORS,    exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# PRODUCT LOOKUP  — code → oz per unit, with fallback by name
# ─────────────────────────────────────────────────────────────────────────────

PRODUCT_OZ: dict[str, float] = {}   # code → oz/unit
PRODUCT_VAT: dict[str, bool]  = {}  # code → has VAT

for metal_key in ('gold', 'silver'):
    for p in _PRODUCTS.get(metal_key, []):
        code = p['code'].strip()
        PRODUCT_OZ[code.lower()]  = float(p['oz'])
        PRODUCT_VAT[code.lower()] = bool(p.get('vat', False))

# ─────────────────────────────────────────────────────────────────────────────
# SOURCE CHANNEL LOOKUP
# ─────────────────────────────────────────────────────────────────────────────

_SOURCE_LOOKUP: dict[str, dict] = {
    k.lower().strip(): v
    for k, v in _PRODUCTS.get('source_channels', {}).items()
}
# Additional digital keywords for fallback
_DIGITAL_KEYWORDS = {'goldstore', 'gold store', 'web', 'app', 'online', 'digital', 'website', 'portal'}


def _classify_source(source: str) -> tuple[str, str]:
    """Return (channel, dealer_name) from the Source Channel field."""
    if not source or str(source).strip() == '' or source == 'nan':
        return 'dealer', ''
    s = str(source).strip()
    lookup = _SOURCE_LOOKUP.get(s.lower())
    if lookup:
        channel = lookup['channel']
        dealer  = '' if lookup['is_digital'] else s
        return channel, dealer
    # Fallback: check digital keywords
    if s.lower() in _DIGITAL_KEYWORDS:
        return 'digital', ''
    return 'dealer', s


def _silo_from_movement(movement: str) -> str:
    """Vault/custody movements → custody silo, otherwise retail."""
    if not movement or str(movement).strip() == '' or movement == 'nan':
        return 'retail'
    m = str(movement).lower()
    if 'vault' in m or 'custody' in m:
        return 'custody'
    return 'retail'


# ─────────────────────────────────────────────────────────────────────────────
# TAB DETECTION
# ─────────────────────────────────────────────────────────────────────────────

SKIP_TAB_KEYWORDS = [
    'mastersheet', 'tables', 'data', 'mt only', 'summary',
    'prices', 'spot', 'settings', 'config', 'pivot', 'index',
    'ref', 'rates', 'holiday', 'calendar',
]
DEAL_TAB_KEYWORDS = [
    'dealing', 'bullion', 'gold', 'silver', 'bar', 'coin', 'kr',
]


def _is_deal_tab(name: str) -> bool:
    n = name.lower().strip()
    if any(skip in n for skip in SKIP_TAB_KEYWORDS):
        return False
    return any(kw in n for kw in DEAL_TAB_KEYWORDS)


def _metal_from_tab(name: str) -> str:
    n = name.lower().strip()
    if 'silver' in n or n.startswith('skr'):
        return 'silver'
    return 'gold'


# ─────────────────────────────────────────────────────────────────────────────
# COLUMN LAYOUTS — explicit 0-based column indices per tab type
#
# KR Dealing:     buybacks A–N  (0–13), sales O–AA (14–26)
# Silver Bullion: buybacks A–P  (0–15), sales Q–AE (16–30)
# Gold Excl KRs:  buybacks A–P  (0–15), sales Q–AE (16–30)
# ─────────────────────────────────────────────────────────────────────────────

_KR_LAYOUT = {
    'buy': {
        'deal_date': 0, 'client_name': 1, 'channel_raw': 2,
        'product_name': 3, 'product_code': 4, 'movement': 5,
        'units': 6, 'equiv_oz': 7, 'oz': 8,
        'spot_price_zar': 9, 'margin_pct': 10,
        'deal_value_zar': 13,
    },
    'sell': {
        # No date on sell side — inherited from buy side (col 0)
        'client_name': 14, 'channel_raw': 15,
        'product_name': 16, 'product_code': 17, 'movement': 18,
        'units': 19, 'equiv_oz': 20, 'oz': 21,
        'spot_price_zar': 22, 'margin_pct': 23,
        'deal_value_zar': 26,
    },
}

_STANDARD_LAYOUT = {
    'buy': {
        'deal_date': 0, 'client_name': 1, 'channel_raw': 2,
        'product_name': 3, 'product_code': 4, 'movement': 5,
        'units': 6, 'equiv_oz': 7, 'oz': 8,
        'spot_price_zar': 9, 'margin_pct': 10,
        'deal_value_zar': 15,
    },
    'sell': {
        # No date on sell side — inherited from buy side (col 0)
        'client_name': 16, 'channel_raw': 17,
        'product_name': 18, 'product_code': 19, 'movement': 20,
        'units': 21, 'equiv_oz': 22, 'oz': 23,
        'spot_price_zar': 24, 'margin_pct': 25,
        'deal_value_zar': 30,
    },
}


def _tab_layout(tab_name: str) -> dict:
    """Return the column layout dict for this tab."""
    n = tab_name.lower()
    if 'kr' in n and 'excluding' not in n and 'excl' not in n:
        return _KR_LAYOUT
    return _STANDARD_LAYOUT


# ─────────────────────────────────────────────────────────────────────────────
# ROW COLOUR CLASSIFICATION
# ─────────────────────────────────────────────────────────────────────────────

def _get_cell_rgb(cell) -> str:
    try:
        fill = cell.fill
        if not fill or fill.fill_type != 'solid':
            return ''
        fc = fill.fgColor
        if fc.type == 'rgb':
            return fc.rgb[-6:]
        if fc.type == 'indexed' and fc.indexed not in (0, 64, 65):
            return ''
    except Exception:
        pass
    return ''


def _classify_rgb(rgb6: str) -> str:
    """
    Orange (confirmed): R>180, G<200, B<80    e.g. FFC000 → confirmed
    Goldish-yellow (quote): R>200, G≥200, B<130  e.g. FFD700 → quote/pipeline
    Blue (proof): dominant B                   → proof
    White/empty → '' (skip)
    """
    if not rgb6 or len(rgb6) < 6:
        return ''
    try:
        r = int(rgb6[0:2], 16)
        g = int(rgb6[2:4], 16)
        b = int(rgb6[4:6], 16)
    except ValueError:
        return ''

    if r > 240 and g > 240 and b > 240:
        return ''   # white/near-white → skip
    if r < 20 and g < 20 and b < 20:
        return ''   # near-black → skip

    # Blue → proof
    if b > 130 and b > r + 40:
        return 'proof'

    # Orange (confirmed) — G clearly below 200 distinguishes from goldish-yellow
    # FFC000: R=255, G=192, B=0  — most common Excel "gold" = confirmed deal
    if r > 180 and g < 200 and b < 80:
        return 'confirmed'

    # Goldish-yellow (pipeline) — both R and G high, B low
    # FFD700: R=255, G=215, B=0 / FFCC00: R=255, G=204, B=0
    if r > 200 and g >= 200 and b < 130:
        return 'quote'

    return 'confirmed'   # any other colour → treat as confirmed


def _log_distinct_rgbs(workbook_path: str, sheet_name: str):
    """Debug: print unique RGB values found in the sheet to help diagnose misclassification."""
    try:
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb[sheet_name]
        seen = set()
        for row in ws.iter_rows():
            for cell in row:
                rgb = _get_cell_rgb(cell)
                if rgb and rgb not in seen:
                    r = int(rgb[0:2], 16)
                    g = int(rgb[2:4], 16)
                    b = int(rgb[4:6], 16)
                    seen.add(rgb)
                    print(f"      RGB #{rgb} (R={r},G={g},B={b}) → {_classify_rgb(rgb) or 'skip'}")
        wb.close()
    except Exception:
        pass


def _read_row_statuses(workbook_path: str, sheet_name: str) -> dict:
    """
    Read the fill colour of the first non-empty cell in each row.
    Returns {excel_row_number: status_string}.
    """
    import warnings
    warnings.filterwarnings('ignore')
    wb = None
    try:
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb[sheet_name]
        statuses = {}
        for row in ws.iter_rows():
            row_num = row[0].row
            status  = ''
            for cell in row:
                rgb = _get_cell_rgb(cell)
                s   = _classify_rgb(rgb)
                if s:
                    status = s
                    break
            statuses[row_num] = status
        return statuses
    except Exception as e:
        print(f"    [colour] Could not read colours from '{sheet_name}': {e}")
        return {}
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────────────────────
# SAFE CELL VALUE HELPER
# ─────────────────────────────────────────────────────────────────────────────

def _v(row, idx):
    """Safely get a value from a pandas row by 0-based column index."""
    try:
        v = row.iloc[idx]
        if pd.isna(v):
            return None
        return v
    except (IndexError, ValueError, TypeError):
        return None


def _flt(row, idx, default=None):
    """Safely get a float from a pandas row; returns default if absent or non-numeric."""
    v = _v(row, idx)
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def _str(row, idx) -> str:
    v = _v(row, idx)
    return str(v).strip() if v is not None else ''


# ─────────────────────────────────────────────────────────────────────────────
# ROW-BY-ROW DEAL EXTRACTOR
# ─────────────────────────────────────────────────────────────────────────────

def _extract_side(row, col_map: dict, deal_type: str,
                  deal_date: str, metal: str, entity: str,
                  source_file: str, status: str, product_type: str):
    """
    Extract one deal (buy OR sell) from a pandas row using explicit column indices.
    Returns a deal dict or None if this side is empty/invalid.
    """
    spot  = _flt(row, col_map['spot_price_zar'])
    units = _flt(row, col_map['units'])

    if not spot or not units or spot <= 0 or units <= 0:
        return None

    # oz: prefer Total Ounces column, fall back to Qty × UOM
    oz_direct = _flt(row, col_map['oz'], 0.0)
    equiv_oz  = _flt(row, col_map['equiv_oz'], None)

    if oz_direct and oz_direct > 0:
        oz    = oz_direct
        equiv = oz / units
    elif equiv_oz and equiv_oz > 0:
        oz    = equiv_oz * units
        equiv = equiv_oz
    else:
        # Try product lookup by code or name
        pcode = _str(row, col_map.get('product_code', -1)).lower() if col_map.get('product_code') is not None else ''
        equiv = PRODUCT_OZ.get(pcode, 1.0)
        oz    = equiv * units

    if oz <= 0:
        return None

    # Margin — stored as decimal (0.037) or percent (3.7); normalise to percent
    margin = _flt(row, col_map['margin_pct'], 0.0)
    if margin is None:
        margin = 0.0
    if -1 < margin < 1 and margin != 0:
        margin = margin * 100.0

    product_name = _str(row, col_map['product_name']) if col_map.get('product_name') is not None else ''
    product_code = _str(row, col_map['product_code']) if col_map.get('product_code') is not None else ''
    channel_raw  = _str(row, col_map['channel_raw'])  if col_map.get('channel_raw')  is not None else ''
    movement     = _str(row, col_map['movement'])      if col_map.get('movement')     is not None else ''

    channel, dealer_name = _classify_source(channel_raw)
    silo                 = _silo_from_movement(movement)

    return {
        'entity':         entity,
        'metal':          metal,
        'deal_type':      deal_type,
        'deal_date':      deal_date,
        'dealer_name':    dealer_name,
        'silo':           silo,
        'channel':        channel,
        'product_code':   product_code,
        'product_name':   product_name,
        'units':          units,
        'equiv_oz':       equiv,
        'oz':             oz,
        'spot_price_zar': spot,
        'margin_pct':     margin,
        'source_file':    source_file,
        'status':         status,
        'product_type':   product_type,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PARSE ONE TAB — row-by-row, both sides share the row date
# ─────────────────────────────────────────────────────────────────────────────

def _parse_deal_tab(df: pd.DataFrame, metal: str, entity: str,
                    source_file: str, row_statuses: dict,
                    tab_name: str) -> list:
    """
    Parse all deals from one sheet tab.

    Each row may contain a BUY deal (left columns) AND/OR a SELL deal
    (right columns). Both sides share the date from column 0 (buy side).
    Row colour determines status: confirmed / quote / proof.
    """
    layout     = _tab_layout(tab_name)
    buy_cols   = layout['buy']
    sell_cols  = layout['sell']

    has_colours = bool(row_statuses) and any(s for s in row_statuses.values())
    deals       = []
    last_date   = date.today().isoformat()

    for pandas_idx, row in df.iterrows():
        # ── colour / status ───────────────────────────────────────────────
        if has_colours:
            excel_row = pandas_idx + 2   # header is row 1, data starts row 2
            status    = row_statuses.get(excel_row, '')
            if not status:
                continue   # no colour → blank / header → skip
        else:
            status = 'confirmed'

        db_status    = 'confirmed' if status in ('confirmed', 'proof') else 'quote'
        product_type = 'proof'    if status == 'proof'                 else 'bullion'

        # ── date — from buy-side col 0, forward-filled ────────────────────
        raw_date = _v(row, buy_cols['deal_date'])
        if raw_date is not None:
            try:
                last_date = pd.to_datetime(raw_date).date().isoformat()
            except Exception:
                pass
        deal_date = last_date

        # ── extract buy side ──────────────────────────────────────────────
        buy = _extract_side(row, buy_cols, 'buy', deal_date, metal, entity,
                            source_file, db_status, product_type)
        if buy:
            deals.append(buy)

        # ── extract sell side ─────────────────────────────────────────────
        sell = _extract_side(row, sell_cols, 'sell', deal_date, metal, entity,
                             source_file, db_status, product_type)
        if sell:
            deals.append(sell)

    return deals


# ─────────────────────────────────────────────────────────────────────────────
# SHEET NAME HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _get_sheet_names(filepath: str) -> list:
    """Fast sheet name extraction via zipfile (no formula parsing)."""
    import zipfile, xml.etree.ElementTree as ET
    try:
        with zipfile.ZipFile(filepath) as z:
            with z.open('xl/workbook.xml') as f:
                tree = ET.parse(f)
                ns   = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                return [s.get('name', '') for s in tree.findall('.//ns:sheet', ns)]
    except Exception:
        wb    = load_workbook(filepath, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def process_file(filepath: str, entity: str = 'SABIS',
                 display_name: str = None) -> dict:
    """
    Process a dealer Excel file. Auto-detects SABIS format vs generic.
    Returns a summary dict with status, deal counts, warnings.
    """
    import warnings
    warnings.filterwarnings('ignore', category=UserWarning)

    filename = display_name or os.path.basename(filepath)
    print(f"\nProcessing: {filename}")

    # Copy to local temp only if not already there.
    # Compare normalised paths to avoid Windows case-mismatch WinError 32.
    src_abs  = os.path.normcase(os.path.abspath(filepath))
    dst_path = os.path.normcase(
        os.path.abspath(os.path.join(tempfile.gettempdir(), os.path.basename(filepath)))
    )
    already_at_dst = (src_abs == dst_path)

    if already_at_dst:
        tmp_path = filepath
    else:
        tmp_path = os.path.join(tempfile.gettempdir(), os.path.basename(filepath))
        try:
            shutil.copy2(filepath, tmp_path)
        except Exception as e:
            _move_to_errors(filepath, f"Could not read file: {e}")
            return {'status': 'error', 'file': filename, 'error': str(e)}

    try:
        sheet_names = _get_sheet_names(tmp_path)
    except Exception as e:
        if not already_at_dst:
            _move_to_errors(filepath, str(e))
        return {'status': 'error', 'file': filename, 'error': str(e)}

    deal_tabs = [s for s in sheet_names if _is_deal_tab(s)]

    if deal_tabs:
        return _process_sabis_file(filepath, filename, deal_tabs, entity, tmp_path)
    return _process_generic_file(filepath, filename, tmp_path)


def _process_sabis_file(filepath: str, filename: str,
                        deal_tabs: list, entity: str,
                        read_from: str = None) -> dict:
    """Process an SABIS multi-tab dealing workbook."""
    src            = read_from or filepath
    all_errors     = []
    processed      = []
    pipeline_added = []

    print(f"  Deal tabs: {deal_tabs}")

    for tab in deal_tabs:
        metal = _metal_from_tab(tab)
        print(f"  Tab: '{tab}' -> metal={metal}")

        row_statuses = _read_row_statuses(src, tab)
        counts = {}
        for s in row_statuses.values():
            counts[s or 'none'] = counts.get(s or 'none', 0) + 1
        print(f"    Row colours: {counts}")
        _log_distinct_rgbs(src, tab)

        try:
            df = pd.read_excel(src, sheet_name=tab, dtype=str, header=0, engine='openpyxl')
            df = df.copy()
        except Exception as e:
            all_errors.append(f"Tab '{tab}': read error — {e}")
            continue

        deals     = _parse_deal_tab(df, metal, entity, filename, row_statuses, tab)
        confirmed = [d for d in deals if d['status'] == 'confirmed']
        quotes    = [d for d in deals if d['status'] == 'quote']
        print(f"    Parsed — confirmed: {len(confirmed)}  quotes: {len(quotes)}")

        for deal in confirmed:
            try:
                result = _process_row(deal, filename)
                if result:
                    processed.append(result)
            except Exception as e:
                all_errors.append(f"Tab '{tab}': row error — {e}")

        for deal in quotes:
            try:
                _insert_pipeline_row(deal)
                pipeline_added.append(deal)
            except Exception as e:
                all_errors.append(f"Tab '{tab}': pipeline error — {e}")

    if not processed and not pipeline_added:
        reason = '\n'.join(all_errors) if all_errors else (
            'No coloured deal rows found. Check orange/yellow/blue row highlighting.')
        _move_to_errors(filepath, reason)
        return {'status': 'error', 'file': filename, 'errors': all_errors}

    metal_breakdown = {}
    for d in processed:
        m = d.get('metal', 'unknown')
        metal_breakdown[m] = metal_breakdown.get(m, 0) + 1

    _move_to_processed(filepath)
    print(f"  Imported {len(processed)} confirmed deals {metal_breakdown}, "
          f"{len(pipeline_added)} pipeline. Warnings: {len(all_errors)}")
    return {
        'status':            'success' if not all_errors else 'partial',
        'file':              filename,
        'deals_imported':    len(processed),
        'pipeline_imported': len(pipeline_added),
        'metal_breakdown':   metal_breakdown,
        'warnings':          all_errors,
        'deals':             processed,
    }


def _insert_pipeline_row(deal: dict):
    oz = deal.get('units', 0) * deal.get('equiv_oz', 1.0)
    insert_pipeline({
        'entity':         deal['entity'],
        'metal':          deal['metal'],
        'deal_type':      deal['deal_type'],
        'deal_date':      deal.get('deal_date', ''),
        'client_name':    deal.get('dealer_name', ''),
        'product_name':   deal.get('product_name', ''),
        'product_code':   deal.get('product_code', ''),
        'units':          deal.get('units', 0),
        'oz':             oz,
        'spot_price_zar': deal.get('spot_price_zar', 0),
        'margin_pct':     deal.get('margin_pct', 0),
        'deal_value_zar': deal.get('spot_price_zar', 0) * oz,
        'product_type':   deal.get('product_type', 'bullion'),
        'source_file':    deal.get('source_file', ''),
    })


# ─────────────────────────────────────────────────────────────────────────────
# GENERIC FORMAT (fallback for non-SABIS sheets)
# ─────────────────────────────────────────────────────────────────────────────

COLUMN_MAP = {
    'deal_type':      ['type', 'deal type', 'buy/sell'],
    'deal_date':      ['date', 'deal date'],
    'entity':         ['entity', 'company'],
    'metal':          ['metal', 'commodity'],
    'silo':           ['silo', 'client type'],
    'channel':        ['channel', 'source'],
    'dealer_name':    ['dealer', 'dealer name'],
    'product_code':   ['product code', 'code'],
    'product_name':   ['product', 'product name'],
    'units':          ['units', 'qty', 'quantity'],
    'equiv_oz':       ['equivalent oz', 'oz per unit', 'uom'],
    'spot_price_zar': ['spot', 'spot price'],
    'margin_pct':     ['margin', 'margin %'],
}


def _map_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename    = {}
    lower_map = {c.lower().strip(): c for c in df.columns}
    for internal, variants in COLUMN_MAP.items():
        for v in variants:
            if v.lower() in lower_map:
                rename[lower_map[v.lower()]] = internal
                break
    return df.rename(columns=rename)


def _process_generic_file(filepath: str, filename: str, read_from: str = None) -> dict:
    src = read_from or filepath
    try:
        df = pd.read_excel(src, dtype=str)
    except Exception as e:
        _move_to_errors(filepath, str(e))
        return {'status': 'error', 'file': filename, 'error': str(e)}

    df = _map_columns(df)
    all_errors, processed = [], []

    for i, row in df.iterrows():
        row_dict = row.to_dict()
        try:
            result = _process_row(row_dict, filename)
            if result:
                processed.append(result)
        except Exception as e:
            all_errors.append(f"Row {i+2}: {e}")

    if all_errors and not processed:
        _move_to_errors(filepath, '\n'.join(all_errors))
        return {'status': 'error', 'file': filename, 'errors': all_errors}

    _move_to_processed(filepath)
    return {
        'status':         'success' if not all_errors else 'partial',
        'file':           filename,
        'deals_imported': len(processed),
        'warnings':       all_errors,
        'deals':          processed,
    }


# ─────────────────────────────────────────────────────────────────────────────
# PROCESS A SINGLE DEAL ROW → DB
# ─────────────────────────────────────────────────────────────────────────────

def _process_row(row: dict, source_file: str):
    """Calculate all fields and write one deal to the database."""
    entity     = str(row['entity']).upper()
    metal      = str(row['metal']).lower()
    deal_type  = str(row['deal_type']).lower()
    silo       = str(row.get('silo', 'retail')).lower()
    channel    = str(row.get('channel', 'dealer')).lower()
    units      = float(row['units'])
    spot       = float(row['spot_price_zar'])
    margin_pct = float(row['margin_pct'])

    equiv_oz = float(row.get('equiv_oz', 1.0) or 1.0)
    oz       = float(row.get('oz', 0) or 0) or (equiv_oz * units)

    deal_date = str(row.get('deal_date', date.today().isoformat()))[:10]

    current_inventory = get_inventory(entity, metal)
    provision         = get_provision_mode(current_inventory, metal)
    provision_pct     = provision['rate_pct']

    if deal_type == 'sell':
        effective_price = spot * (1 + margin_pct / 100)
    else:
        effective_price = spot * (1 - margin_pct / 100)

    deal_value = effective_price * oz
    notional   = spot * oz
    mvp        = margin_vs_provision(margin_pct, provision_pct, deal_type)
    gp         = calc_gp_contribution(mvp['profit_pct'], notional)

    inv_delta     = oz if deal_type == 'buy' else -oz
    new_inventory = current_inventory + inv_delta
    flipped       = (current_inventory >= 0 and new_inventory < 0) or \
                    (current_inventory < 0  and new_inventory >= 0)

    deal_record = {
        'entity':              entity,
        'metal':               metal,
        'deal_type':           deal_type,
        'deal_date':           deal_date,
        'dealer_name':         str(row.get('dealer_name', '')),
        'silo':                silo,
        'channel':             channel,
        'product_code':        str(row.get('product_code', '')),
        'product_name':        str(row.get('product_name', '')),
        'equiv_oz_per_unit':   equiv_oz,
        'units':               units,
        'oz':                  oz,
        'spot_price_zar':      spot,
        'margin_pct':          margin_pct,
        'effective_price_zar': effective_price,
        'deal_value_zar':      deal_value,
        'provision_pct':       provision_pct,
        'profit_margin_pct':   mvp['profit_pct'],
        'gp_contribution_zar': gp,
        'inventory_after_oz':  new_inventory,
        'provision_flipped':   1 if flipped else 0,
        'source_file':         source_file,
        'status':              str(row.get('status', 'confirmed')),
        'product_type':        str(row.get('product_type', 'bullion')),
    }

    deal_id = insert_deal(deal_record)
    if deal_id < 0:
        return None   # duplicate — silently skip

    update_inventory(entity, metal, inv_delta)
    insert_cash_flow({
        'entity':      entity,
        'flow_date':   deal_date,
        'flow_type':   'deal',
        'direction':   'in' if deal_type == 'sell' else 'out',
        'amount_zar':  deal_value,
        'description': f"{deal_type.title()} {oz:.4f}oz {metal} @ {spot} spot {margin_pct:+.2f}%",
        'deal_id':     deal_id,
    })

    deal_record['deal_id'] = deal_id
    return deal_record


# ─────────────────────────────────────────────────────────────────────────────
# FILE MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

def _move_to_processed(filepath: str):
    dest = os.path.join(PROCESSED, os.path.basename(filepath))
    try:
        shutil.move(filepath, dest)
        print(f"  → Moved to processed: {dest}")
    except Exception:
        pass   # temp file already cleaned up by upload endpoint


def _move_to_errors(filepath: str, reason: str):
    dest = os.path.join(ERRORS, os.path.basename(filepath))
    try:
        shutil.move(filepath, dest)
        with open(dest + '.error.txt', 'w') as f:
            f.write(reason)
        print(f"  ✗ Moved to errors. Reason: {reason[:120]}")
    except Exception:
        pass
